"""
資料匯入通用工具模組
提供從 Excel 和 CSV 檔案匯入各種資料類型的功能
"""

import csv
import openpyxl
import re
from decimal import Decimal
from datetime import datetime
from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone
from crm.models import Owner, Project, Category
from users.models import UserProfile
import chardet


class ImportResult:
    """匯入結果類別"""
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors = []
        self.warnings = []
    
    def add_error(self, row_number, message):
        """新增錯誤訊息"""
        self.error_count += 1
        self.errors.append({
            'row': row_number,
            'message': message
        })
    
    def add_warning(self, row_number, message):
        """新增警告訊息"""
        self.warnings.append({
            'row': row_number,
            'message': message
        })
    
    def add_success(self):
        """增加成功計數"""
        self.success_count += 1
    
    def to_dict(self):
        """轉換為字典格式"""
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'total_processed': self.success_count + self.error_count,
            'errors': self.errors,
            'warnings': self.warnings
        }


class BaseImporter:
    """基礎匯入類別"""
    
    def __init__(self):
        self.result = ImportResult()
    
    def detect_encoding(self, file_path):
        """偵測檔案編碼"""
        try:
            with open(file_path, 'rb') as file:
                raw_data = file.read(10000)  # 讀取前10000個位元組來偵測
                result = chardet.detect(raw_data)
                encoding = result['encoding']
                # 如果偵測到BIG5或CP950,統一使用cp950(相容性更好)
                if encoding and encoding.upper() in ['BIG5', 'CP950']:
                    return 'cp950'
                # 如果偵測到UTF-8 with BOM
                if encoding and encoding.upper() in ['UTF-8-SIG', 'UTF-8']:
                    return 'utf-8-sig'
                return encoding if encoding else 'utf-8-sig'
        except:
            # 如果偵測失敗,嘗試常見的編碼順序
            return 'utf-8-sig'
    
    def parse_date(self, date_value):
        """解析日期值"""
        if not date_value:
            return None
        
        if isinstance(date_value, datetime):
            return date_value.date()
        
        if isinstance(date_value, str):
            # 嘗試不同的日期格式
            date_formats = [
                '%Y/%m/%d', '%Y-%m-%d', '%Y年%m月%d日',
                '%m/%d/%Y', '%d/%m/%Y'
            ]
            
            for fmt in date_formats:
                try:
                    return datetime.strptime(date_value.strip(), fmt).date()
                except ValueError:
                    continue
        
        return None
    
    def parse_boolean(self, value):
        """解析布林值"""
        if isinstance(value, bool):
            return value
        
        if isinstance(value, str):
            value = value.strip().lower()
            return value in ['是', 'true', '1', 'yes', 'y']
        
        return bool(value)
    
    def parse_decimal(self, value):
        """解析數值"""
        if not value:
            return Decimal('0')
        
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        if isinstance(value, str):
            # 移除非數字字元（保留小數點和負號）
            cleaned = ''.join(ch for ch in value if ch.isdigit() or ch in '.-')
            try:
                return Decimal(cleaned) if cleaned else Decimal('0')
            except:
                return Decimal('0')
        
        return Decimal('0')
    
    def safe_get_value(self, row, index, default=''):
        """安全取得列表中的值"""
        try:
            return row[index] if index < len(row) and row[index] is not None else default
        except (IndexError, TypeError):
            return default


class OwnerImporter(BaseImporter):
    """業主資料匯入器"""
    
    def import_from_excel(self, file_path):
        """從 Excel 檔案匯入業主資料"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            if not workbook.sheetnames:
                self.result.add_error(0, "Excel 檔案中沒有工作表")
                return self.result
            
            # 使用第一張工作表
            sheet = workbook[workbook.sheetnames[0]]
            
            with transaction.atomic():
                # 跳過標題列，從第二列開始
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        self._process_owner_row(row, row_num)
                    except Exception as e:
                        self.result.add_error(row_num, f"處理資料時發生錯誤: {str(e)}")
            
        except Exception as e:
            self.result.add_error(0, f"讀取 Excel 檔案時發生錯誤: {str(e)}")
        
        return self.result
    
    def import_from_csv(self, file_path, encoding=None):
        """從 CSV 檔案匯入業主資料"""
        try:
            # 自動偵測編碼
            if encoding is None:
                encoding = self.detect_encoding(file_path)
            
            with open(file_path, 'r', encoding=encoding, errors='replace') as file:
                # 使用 Sniffer 自動偵測分隔符號
                sample = file.read(1024)
                file.seek(0)
                try:
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter
                except:
                    # 如果偵測失敗，預設使用逗號
                    delimiter = ','
                
                reader = csv.reader(file, delimiter=delimiter)
                next(reader)  # 跳過標題列
                
                with transaction.atomic():
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            self._process_owner_row(row, row_num)
                        except Exception as e:
                            self.result.add_error(row_num, f"處理資料時發生錯誤: {str(e)}")
                            
        except Exception as e:
            self.result.add_error(0, f"讀取 CSV 檔案時發生錯誤: {str(e)}")
        
        return self.result
    
    def _process_owner_row(self, row, row_num):
        """處理單筆業主資料"""
        # 預期欄位：公司名稱, 電話, 傳真, 電子郵件, 手機, 地址, 聯絡人
        company_name = self.safe_get_value(row, 0)
        phone = self.safe_get_value(row, 1)
        fax = self.safe_get_value(row, 2)
        email = self.safe_get_value(row, 3)
        mobile = self.safe_get_value(row, 4)
        address = self.safe_get_value(row, 5)
        contact_person = self.safe_get_value(row, 6)
        
        # 驗證必要欄位
        if not company_name:
            self.result.add_error(row_num, "公司名稱不能為空")
            return
        
        # 檢查是否已存在相同公司名稱
        if Owner.objects.filter(company_name=company_name).exists():
            self.result.add_warning(row_num, f"公司 '{company_name}' 已存在，跳過匯入")
            return
          # 建立業主資料
        owner = Owner.objects.create(
            company_name=company_name,
            tax_id="",  # 暫時設為空
            phone=phone or "待修改-phone",
            fax=fax or "待修改-fax",
            email=email or "待修改-email",
            mobile=mobile or "待修改-mobile",
            address=address or "待修改-address",
            contact_person=contact_person or "待修改-contact_person",
        )
        
        # 更新 tax_id 為資料庫生成的 id，確保不超過 10 個字符
        # 使用 "待修" + id 的格式，如果超過 10 字符則截斷
        temp_tax_id = f"待修{owner.id}"
        if len(temp_tax_id) > 10:
            # 如果超過 10 字符，只保留 id 的部分並加上前綴
            id_str = str(owner.id)
            max_id_length = 10 - 2  # 為 "待修" 預留 2 個字符（中文字符）的空間  
            if len(id_str) <= max_id_length:
                temp_tax_id = f"待修{id_str}"
            else:
                # 如果 id 太長，直接使用 id 的後幾位
                temp_tax_id = id_str[-10:]
        
        owner.tax_id = temp_tax_id
        owner.save(update_fields=["tax_id"])
        
        self.result.add_success()


class EmployeeImporter(BaseImporter):
    """員工帳號匯入器"""
    
    def import_from_excel(self, file_path):
        """從 Excel 檔案匯入員工帳號資料"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            if not workbook.sheetnames:
                self.result.add_error(0, "Excel 檔案中沒有工作表")
                return self.result
            
            # 使用第一張工作表
            sheet = workbook[workbook.sheetnames[0]]
            
            with transaction.atomic():
                # 跳過標題列，從第二列開始
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        self._process_employee_row(row, row_num)
                    except Exception as e:
                        self.result.add_error(row_num, f"處理資料時發生錯誤: {str(e)}")
            
        except Exception as e:
            self.result.add_error(0, f"讀取 Excel 檔案時發生錯誤: {str(e)}")
        
        return self.result
    
    def _process_employee_row(self, row, row_num):
        """處理單筆員工資料"""
        # 預期欄位：A欄=員編, B欄=姓名
        employee_id = self.safe_get_value(row, 0)
        name = self.safe_get_value(row, 1)
        
        # 轉換為字串並清理
        employee_id = str(employee_id).strip() if employee_id else ""
        name = str(name).strip() if name else ""
        
        # 驗證必要欄位（員編和姓名都不能為空）
        if not employee_id or not name:
            self.result.add_warning(row_num, "員編或姓名為空，跳過此筆資料")
            return
        
        # 檢查是否已存在相同姓名的使用者
        if User.objects.filter(username=name).exists():
            self.result.add_warning(row_num, f"帳號 '{name}' 已存在，跳過匯入")
            return

        # 拆分姓名為 first_name 和 last_name
        if len(name) > 1:
            first_name = name[0]
            last_name = name[1:]
        else:
            first_name = name
            last_name = ""
        
        # 建立使用者帳號，username 設為完整中文姓名
        user = User.objects.create(
            username=name,  # username 改為完整中文姓名
            first_name=first_name,
            last_name=last_name,
        )
        
        # 設定密碼
        user.set_password("12345678")
        user.save()
        
        # 建立 UserProfile
        UserProfile.objects.create(
            user=user,
            name=name,
            is_project_manager=True
        )
        
        self.result.add_success()


class ProjectImporter(BaseImporter):
    """專案資料匯入器"""
    
    def import_from_excel(self, file_path):
        """從 Excel 檔案匯入專案資料"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            if not workbook.sheetnames:
                self.result.add_error(0, "Excel 檔案中沒有工作表")
                return self.result
            
            # 使用第一張工作表
            sheet = workbook[workbook.sheetnames[0]]
            column_mapping = self._map_excel_columns(sheet)
            
            if not column_mapping:
                self.result.add_error(0, "找不到有效的欄位映射")
                return self.result
            
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        self._process_project_row(row, row_num, column_mapping)
                    except Exception as e:
                        self.result.add_error(row_num, f"處理資料時發生錯誤: {str(e)}")
            
        except Exception as e:
            self.result.add_error(0, f"讀取 Excel 檔案時發生錯誤: {str(e)}")
        
        return self.result
    
    def import_from_csv(self, file_path, encoding=None):
        """從 CSV 檔案匯入專案資料"""
        try:
            # 自動偵測編碼
            if encoding is None:
                encoding = self.detect_encoding(file_path)
            
            with open(file_path, 'r', encoding=encoding, errors='replace') as file:
                # 使用 Sniffer 自動偵測分隔符號
                sample = file.read(1024)
                file.seek(0)
                try:
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter
                except:
                    # 如果偵測失敗,預設使用逗號
                    delimiter = ','
                
                reader = csv.reader(file, delimiter=delimiter)
                headers = next(reader)  # 讀取標題列
                
                # 映射 CSV 欄位
                column_mapping = self._map_csv_columns(headers)
                
                if not column_mapping:
                    self.result.add_error(0, "找不到有效的欄位映射")
                    return self.result
                
                with transaction.atomic():
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            self._process_project_row(row, row_num, column_mapping)
                        except Exception as e:
                            self.result.add_error(row_num, f"處理資料時發生錯誤: {str(e)}")
                            
        except Exception as e:
            self.result.add_error(0, f"讀取 CSV 檔案時發生錯誤: {str(e)}")
        
        return self.result
    
    def _map_excel_columns(self, sheet):
        """映射 Excel 欄位"""
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        column_mapping = {
            "案件類別": "category_code",
            "年份": "year",
            "案件編號": "project_number",
            "案件負責人": "manager_name",
            "繪圖": "drawing_name",
            "報價": "quotation",
            "案件名稱": "project_name",
            "業主": "owner_name",
            "聯絡方式": "contact_info",
            "是否完成": "is_completed",
            "收款日期": "payment_date",
            "是否請款": "is_invoiced",
            "請款日期": "invoice_date",
            "請款金額": "invoice_amount",
            "請款備註": "invoice_notes",
            "發票日期": "invoice_issue_date",
            "是否收款": "is_paid",
            "備註": "notes",
        }
        return {
            column_mapping.get(header, None): idx 
            for idx, header in enumerate(headers) 
            if header in column_mapping
        }
    
    def _map_csv_columns(self, headers):
        """映射 CSV 欄位"""
        column_mapping = {
            "案件類別": "category_code",
            "年份": "year",
            "案件編號": "project_number",
            "案件負責人": "manager_name",
            "繪圖": "drawing_name",
            "報價": "quotation",
            "案件名稱": "project_name",
            "業主": "owner_name",
            "聯絡方式": "contact_info",
            "是否完成": "is_completed",
            "收款日期": "payment_date",
            "是否請款": "is_invoiced",
            "請款日期": "invoice_date",
            "請款金額": "invoice_amount",
            "請款備註": "invoice_notes",
            "發票日期": "invoice_issue_date",
            "是否收款": "is_paid",
            "備註": "notes",
        }
        
        return {
            column_mapping.get(header, None): idx 
            for idx, header in enumerate(headers) 
            if header in column_mapping
        }
        
    def _process_project_row(self, row, row_num, column_mapping):
        """處理單筆專案資料 - 參考 import_projects_from_excel 的邏輯"""
        try:
            # 根據欄位名稱映射提取資料
            data = {key: self.safe_get_value(row, idx) for key, idx in column_mapping.items() if idx < len(row)}

            # 檢查必要欄位是否存在
            if not data.get("category_code") or not data.get("year") or not data.get("project_number"):
                self.result.add_error(row_num, f"跳過資料行，缺少必要欄位: 案件類別={data.get('category_code')}, 年份={data.get('year')}, 案件編號={data.get('project_number')}")
                return

            # 處理資料並新增專案
            year = int(data["year"]) if data["year"] else None
            project_number = str(data["project_number"]).zfill(3) if data["project_number"] else None
            quotation = self.parse_decimal(data.get("quotation"))
            invoice_amount = self.parse_decimal(data.get("invoice_amount"))
            # 將 is_completed 布林值轉換為 status
            is_completed_bool = self.parse_boolean(data.get("is_completed"))
            status = 'completed' if is_completed_bool else 'in_progress'
            is_invoiced = self.parse_boolean(data.get("is_invoiced"))
            is_paid = self.parse_boolean(data.get("is_paid"))
            payment_date = self.parse_date(data.get("payment_date"))
            invoice_date = self.parse_date(data.get("invoice_date"))
            invoice_issue_date = self.parse_date(data.get("invoice_issue_date"))

            # 查找或新增類別
            # 解析類別代碼和描述
            category_data = data["category_code"]
            # 修改正規表達式,支援英文字母加數字的組合 (如 E01, A1, BC23 等)
            match = re.match(r'^([A-Za-z]+\d*)\s*(.*)$', category_data)
            if match:
                code = match.group(1)
                description = match.group(2) or "此類別從excel自動新增"
            else:
                code = category_data
                description = "此類別從excel自動新增"
            category, _ = Category.objects.get_or_create(
                code=code,
                defaults={"description": description}
            )
            
            # 查找或新增業主
            owner_name = data.get("owner_name", "").strip()
            if not owner_name:
                self.result.add_error(row_num, "業主名稱不能為空")
                return
                
            owner_queryset = Owner.objects.filter(company_name=owner_name)
            if owner_queryset.count() > 2:
                self.result.add_warning(row_num, f"業主 '{owner_name}' 超過2筆，跳過新增")
                return
            elif owner_queryset.exists():
                owner = owner_queryset.first()
            else:
                # 先創建業主，暫時設 tax_id 為空
                owner = Owner.objects.create(
                    company_name=owner_name,
                    phone="待修改-phone",
                    fax="待修改-fax",
                    email="待修改-email",
                    mobile="待修改-mobile",
                    address="待修改-address",
                    contact_person="待修改-contact_person",
                )
                
                # 使用與 OwnerImporter 相同的邏輯來設定 tax_id
                temp_tax_id = f"待修{owner.id}"
                if len(temp_tax_id) > 10:
                    # 如果超過 10 字符，只保留 id 的部分並加上前綴
                    id_str = str(owner.id)
                    max_id_length = 10 - 2  # 為 "待修" 預留 2 個字符的空間  
                    if len(id_str) <= max_id_length:
                        temp_tax_id = f"待修{id_str}"
                    else:
                        # 如果 id 太長，直接使用 id 的後幾位
                        temp_tax_id = id_str[-10:]
                
                owner.tax_id = temp_tax_id
                owner.save(update_fields=["tax_id"])
                self.result.add_warning(row_num, f"自動建立業主: {owner_name}")

            # 查找或新增案件負責人
            managers = []
            if data.get("manager_name"):
                # 處理多個負責人情況：支援逗號、換行符、空格分隔
                for name in re.split(r'[,\n\s]+', data["manager_name"]):
                    name = name.strip()
                    if not name:
                        continue
                        
                    # 先嘗試用完整姓名(username)查找
                    manager = User.objects.filter(username=name).first()
                    
                    if not manager:
                        # 如果找不到,嘗試用 first_name + last_name 查找
                        if len(name) > 1:
                            first_name = name[0]
                            last_name = name[1:]
                        else:
                            first_name = name
                            last_name = ""
                        
                        manager = User.objects.filter(
                            first_name=first_name, 
                            last_name=last_name
                        ).first()
                    
                    if not manager:
                        # 如果還是找不到,創建新使用者
                        if len(name) > 1:
                            first_name = name[0]
                            last_name = name[1:]
                        else:
                            first_name = name
                            last_name = ""
                        
                        manager = User.objects.create(
                            username=name,  # username 設為完整中文姓名
                            first_name=first_name,
                            last_name=last_name,
                        )
                        
                        # 設定密碼
                        manager.set_password("12345678")
                        manager.save()
                        
                        # 創建 UserProfile
                        UserProfile.objects.create(
                            user=manager, 
                            name=name, 
                            is_project_manager=True
                        )
                    
                    # 無論是新使用者或現有使用者,都加入 managers 列表以便後續關聯專案
                    managers.append(manager)

            # 檢查是否已存在專案
            if Project.objects.filter(year=year, category=category, project_number=project_number).exists():
                self.result.add_warning(row_num, f"跳過重複專案: {year}-{code}-{project_number}")
                return

            # 新增專案
            project = Project.objects.create(
                owner=owner,
                year=year,
                project_number=project_number,
                name=data.get("project_name") or "",
                drawing=data.get("drawing_name") or "",
                contact_info=data.get("contact_info") or "",
                notes=data.get("notes") or "",
                status=status,  # 使用 status 而不是 is_completed
                category=category,
                is_invoiced=is_invoiced,
                invoice_date=invoice_date,
                invoice_amount=invoice_amount,
                payment_date=payment_date,
                invoice_issue_date=invoice_issue_date,
                invoice_notes=data.get("invoice_notes") or "",
                is_paid=is_paid,
            )

            # 設定案件負責人
            if managers:
                project.managers.set(managers)

            self.result.add_success()

        except Exception as e:
            self.result.add_error(row_num, f"處理專案資料時發生錯誤: {str(e)}")


def import_owners_from_file(file_path, file_type='excel'):
    """業主資料匯入的統一入口點"""
    importer = OwnerImporter()
    
    if file_type.lower() == 'excel':
        return importer.import_from_excel(file_path)
    elif file_type.lower() == 'csv':
        return importer.import_from_csv(file_path)
    else:
        result = ImportResult()
        result.add_error(0, f"不支援的檔案類型: {file_type}")
        return result


def import_projects_from_file(file_path, file_type='excel'):
    """專案資料匯入的統一入口點"""
    importer = ProjectImporter()
    
    if file_type.lower() == 'excel':
        return importer.import_from_excel(file_path)
    elif file_type.lower() == 'csv':
        return importer.import_from_csv(file_path)
    else:
        result = ImportResult()
        result.add_error(0, f"不支援的檔案類型: {file_type}")
        return result


def import_employees_from_file(file_path, file_type='excel'):
    """員工帳號匯入的統一入口點"""
    importer = EmployeeImporter()
    
    if file_type.lower() == 'excel':
        return importer.import_from_excel(file_path)
    else:
        result = ImportResult()
        result.add_error(0, f"不支援的檔案類型: {file_type}")
        return result
