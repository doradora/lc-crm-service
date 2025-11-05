import os
import traceback
import logging
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from copy import copy
from datetime import datetime
from django.conf import settings
from crm.models import ProjectInvoice  # 檔案最上方已經 import


logger = logging.getLogger(__name__)


def generate_payment_excel(payment):
    """
    產生請款單 Excel，回傳 BytesIO 物件
    Args:
        payment: Payment instance
    Returns:
        BytesIO: Excel 檔案內容
    Raises:
        FileNotFoundError: 樣板檔案不存在
        ValueError: payment 關聯資料不完整
        Exception: 其他錯誤
    """
    template_file = os.path.join(settings.BASE_DIR, "crm", "excel", "payment.xlsx")
    if not os.path.exists(template_file):
        logger.error(f"Excel 樣板檔案不存在: {template_file}")
        raise FileNotFoundError(f"Excel 樣板檔案不存在: {template_file}")
    try:
        wb = load_workbook(template_file)
        if "請款單" not in wb.sheetnames:
            logger.error(f"Excel 樣板缺少 '請款單' 工作表: {template_file}")
            raise ValueError(f"Excel 樣板缺少 '請款單' 工作表: {template_file}")
        original_ws = wb["請款單"]

        # 檢查 payment 關聯資料
        if not payment.company:
            logger.error(f"請款單缺少公司資訊 (payment id={payment.id})")
            raise ValueError("請款單缺少公司資訊")
        if not payment.owner:
            logger.error(f"請款單缺少業主資訊 (payment id={payment.id})")
            raise ValueError("請款單缺少業主資訊")

        # 設定收款公司資訊
        company = payment.company
        original_ws["B20"] = f"公司名稱：{company.name}"
        original_ws["B21"] = f"負責人：{company.responsible_person}"
        original_ws["B22"] = f"統一編號：{company.tax_id}"
        original_ws["B23"] = f"地址：{company.address}"
        original_ws["B24"] = f"電話：{company.phone}"
        original_ws["B25"] = f"傳真：{company.fax if company.fax else ''}"
        original_ws["B26"] = f"聯絡人：{company.contact_person}"

        # 設定匯款帳號資訊（如果有的話）
        if payment.selected_bank_account:
            bank_account = payment.selected_bank_account
            original_ws["B29"] = f"戶名：{bank_account.account_name}"
            original_ws["B30"] = f"匯款帳號：{bank_account.account_number}"
            original_ws["B31"] = f"銀行名稱：{bank_account.bank_name}(機構代碼:{bank_account.bank_code})"
            
            # 設定藍色外框
            # 上方儲存格 - 只有上、左、右邊框
            top_border = Border(
                top=Side(style="medium", color="000000"),
                left=Side(style="medium", color="000000"),
                right=Side(style="medium", color="000000")
            )
            # 中間儲存格 - 只有左、右邊框
            middle_border = Border(
                left=Side(style="medium", color="000000"),
                right=Side(style="medium", color="000000")
            )
            # 下方儲存格 - 只有下、左、右邊框
            bottom_border = Border(
                bottom=Side(style="medium", color="000000"),
                left=Side(style="medium", color="000000"),
                right=Side(style="medium", color="000000")
            )
            
            original_ws["B29"].border = top_border
            original_ws["B30"].border = middle_border
            original_ws["B31"].border = bottom_border

        # 獲取請款單關聯的專案明細
        payment_projects = payment.paymentproject_set.select_related("project", "project__category")
        owner = payment.owner.company_name

        # --- 專案排序：先依類別名稱字母順，再依自訂欄位值排序 ---
        def get_sort_key(pp):
            project = pp.project
            if not project or not project.category:
                return (0, '', 0)

            # 年份
            year = project.year if project.year else 0
            # 類別代碼
            category_code = project.category.code if project.category.code else ''
            # 案號轉換為數字進行排序
            project_number = project.project_number
            try:
                project_number_int = int(project_number) if project_number else 0
            except (ValueError, TypeError):
                project_number_int = 0

            return (year, category_code, project_number_int)
        payment_projects = sorted(payment_projects, key=get_sort_key)

        # --- 創建專案明細列表 ---
        project_details = []
        for idx, pp in enumerate(payment_projects, 1):
            if not pp.project:
                logger.warning(f"請款單關聯專案不存在 (payment id={payment.id}, paymentproject id={pp.id})")
                continue
            project = pp.project
            # 工程明細：專案名稱+報告書名稱（換行）
            if getattr(project, "report_name", None):
                detail_name = f"{project.name}\n{project.report_name}"
            else:
                detail_name = project.name
            detail = {
                "項次": idx,
                "工程明細": detail_name,
                "金額": pp.amount,
                "project": project,
            }
            project_details.append(detail)
        if not project_details:
            logger.warning(f"請款單無專案明細 (payment id={payment.id})")
            project_details = [
                {"項次": 1, "工程明細": "無專案明細", "金額": 0, "project": None}
            ]

        total_pages = (len(project_details) + 9) // 10
        # --- 收集所有自訂欄位，依「類別代碼字母順」再依 order 排序，案號最前 ---
        all_custom_fields = []  # [(category_code, field_name, order, display_name)]
        for detail in project_details:
            project = detail["project"]
            if project and project.category and project.category.custom_field_schema:
                category_code = getattr(project.category, "code", "") or ""
                for field_name, field_props in project.category.custom_field_schema.items():
                    display_name = field_props.get("display_name", field_name)
                    order = field_props.get("order", 0)
                    all_custom_fields.append((category_code, field_name, order, display_name))
        # 去除重複 (category_code, field_name)
        seen = set()
        unique_fields = []
        for item in all_custom_fields:
            key = (item[0], item[1])
            if key not in seen:
                seen.add(key)
                unique_fields.append(item)
        # 排序：類別代碼字母順 → order → field_name
        unique_fields.sort(key=lambda x: (x[0], x[2], x[1]))
        # 案號(pcode)永遠最前
        sorted_custom_fields = [("pcode", {"display_name": "案號"})]
        for _, field_name, _, display_name in unique_fields:
            if field_name != "pcode":
                sorted_custom_fields.append((field_name, {"display_name": display_name}))
        # 產生 custom_field_columns
        custom_field_columns = {}
        next_col = 4
        for field_name, field_props in sorted_custom_fields:
            custom_field_columns[field_name] = next_col
            next_col += 1
        max_column = 3
        if custom_field_columns:
            max_column = max(max_column, max(custom_field_columns.values()))

        # 計算所有專案總金額
        total_amount = sum(pp.amount or 0 for pp in payment.paymentproject_set.all())

        # 準備 LOGO 圖片路徑
        logo_path = os.path.join(settings.BASE_DIR, "crm", "excel", "logo.jpg")
        logo_exists = os.path.exists(logo_path)
        if not logo_exists:
            logger.warning(f"LOGO 圖片不存在: {logo_path}")

        for page in range(total_pages):
            if page == 0:
                ws = original_ws
                for col in range(len(sorted_custom_fields)-1):
                    copy_column_and_insert(ws, source_col=5, target_col=4)
            else:
                ws = wb.copy_worksheet(original_ws)
                ws.title = f"請款單-第{page+1}頁"
                for r in range(7, 17):
                    for c in range(1, max_column + 2):
                        cell = ws.cell(row=r, column=c)
                        cell.value = None
            
            # 在每一頁加入 LOGO 圖片
            if logo_exists:
                try:
                    img = Image(logo_path)
                    # 調整圖片大小 (根據附圖位置，放在 B1:B3 左右區域)
                    # 只設定圖片寬度，高度會自動等比例縮放
                    img.width = 120
                    img.height = 67
                    # 將圖片錨定在 B2 儲存格
                    ws.add_image(img, "B2")
                except Exception as e:
                    logger.error(f"插入 LOGO 圖片時發生錯誤: {str(e)}")
            
            ws.print_area = "A1:C34"
            ws.page_margins.left = 0.72
            ws.page_margins.right = 0.72
            ws.page_margins.top = 0.36
            ws.page_margins.bottom = 0.36
            ws.page_margins.header = 0.32
            ws.page_margins.footer = 0.32
            ws["B4"] = payment.payment_number
            # 民國年格式：西元年-1911
            date_obj = payment.date_issued if payment.date_issued else datetime.now()
            roc_year = date_obj.year - 1911
            roc_date_str = f"{roc_year}年{date_obj.month}月{date_obj.day}日"
            ws["C5"] = f"{roc_date_str}"
            ws["B5"] = owner or "未指定業主"
            ws["C17"] = "=SUM(C7:C16)"
            # --- C18顯示所有專案總金額 ---
            ws["C18"] = total_amount
            start_idx = page * 10
            end_idx = min(start_idx + 10, len(project_details))
            page_details = project_details[start_idx:end_idx]
            if sorted_custom_fields:
                header_row = 6
                for field_name, field_props in sorted_custom_fields:
                    col = custom_field_columns[field_name]
                    cell = ws.cell(row=header_row, column=col)
                    cell.value = field_props["display_name"]
                    cell.font = Font(bold=True, size=12, name="Microsoft JhengHei")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            # --- 寫入專案明細 ---
            _fill_project_details(ws, page_details, 7, custom_field_columns)
            # --- 若專案數量<10，於第n+1列工程明細欄填"以下空白" ---
            if len(page_details) < 10:
                empty_row = 7 + len(page_details)
                ws.cell(row=empty_row, column=2).value = "以下空白"
                ws.cell(row=empty_row, column=2).font = Font(size=12, name="Microsoft JhengHei")
                ws.cell(row=empty_row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
                # 金額欄位留空白，不顯示"-"
                ws.cell(row=empty_row, column=3).value = ""
            # --- 7~16列列高45，字體12，微軟正黑體 ---
            for row_idx in range(7, 17):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = Font(size=12, name="Microsoft JhengHei")
                ws.row_dimensions[row_idx].height = 45
            
            # 移除多餘框線 (原金額)
            for row_idx in range(6, 17):
                last_custom_col = max(custom_field_columns.values())
                right_cell = ws.cell(row=row_idx, column=last_custom_col + 1)
                right_cell.border = Border(
                    left=Side(style="mediumDashDot"),
                ) 
                
            for col_idx in range(3, ws.max_column + 1):
                auto_adjust_column_width(ws, col_idx)

        # === 新增：收款註記(發票)區塊填入 ===
        # 收款註記起始列
        receipt_note_start_row = 36
        receipt_note_columns = [
            (1, "收款日"),
            (2, "發票日期/字軌號碼"),
            (3, "案號"),
            (4, "收款方式"),
            (5, "入帳日"),
            (6, "金額"),
            (7, "存放行庫"),
            (8, "備註"),
        ]
        # 寫入標題列
        for col, title in receipt_note_columns:
            cell = original_ws.cell(row=receipt_note_start_row, column=col)
            cell.value = title
            cell.font = Font(size=12, name="Microsoft JhengHei", color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                top=Side(style="mediumDashDot"),
                bottom=Side(style="thin")
            )
        # 寫入發票資料 - 以專案為主體排序
        invoice_qs = payment.invoices.all() if hasattr(payment, 'invoices') else []
        bank_name = payment.selected_bank_account.bank_name if payment.selected_bank_account else ""
        
        # 建立發票與專案的對應關係字典
        project_invoices_map = {}  # {project_id: [(pi, invoice), ...]}
        for invoice in invoice_qs:
            related_projects = ProjectInvoice.objects.filter(invoice=invoice).select_related("project", "project__category")
            for pi in related_projects:
                if pi.project:
                    project_id = pi.project.id
                    if project_id not in project_invoices_map:
                        project_invoices_map[project_id] = []
                    project_invoices_map[project_id].append((pi, invoice))
        
        # 取得所有專案並排序：年份 → 類別代碼 → 案號
        all_projects = []
        for detail in project_details:
            if detail["project"]:
                all_projects.append(detail["project"])
        
        def get_project_sort_key(project):
            if not project:
                return (0, '', 0)
            
            year = project.year if project.year else 0
            category_code = project.category.code if project.category and project.category.code else ''
            try:
                project_number_int = int(project.project_number) if project.project_number else 0
            except (ValueError, TypeError):
                project_number_int = 0
            
            return (year, category_code, project_number_int)
        
        all_projects.sort(key=get_project_sort_key)
        
        row = receipt_note_start_row
        
        # 依專案排序順序寫入資料
        for project in all_projects:
            # 產生案號
            year = getattr(project, "year", None)
            category = getattr(project, "category", None)
            code = getattr(category, "code", "") if category else ""
            number = getattr(project, "project_number", "")
            pcode = ""
            if year and code and number is not None:
                pcode = f"{year}{code}{str(number).zfill(3)}"
            
            # 檢查此專案是否有發票
            project_invoice_data = project_invoices_map.get(project.id, [])
            
            if project_invoice_data:
                # 有發票：寫入每張發票的資料
                for pi, invoice in project_invoice_data:
                    row += 1
                    # 收款日
                    original_ws.cell(row=row, column=1).value = invoice.payment_received_date.strftime('%Y/%m/%d') if invoice.payment_received_date else ""
                    # 發票日期/字軌號碼
                    v = ""
                    if invoice.issue_date:
                        v += invoice.issue_date.strftime('%Y-%m-%d')
                    if invoice.invoice_number:
                        v += f"/{invoice.invoice_number}"
                    original_ws.cell(row=row, column=2).value = v
                    # 案號
                    original_ws.cell(row=row, column=3).value = pcode
                    # 收款方式
                    original_ws.cell(row=row, column=4).value = dict(invoice.PAYMENT_METHOD_CHOICES).get(invoice.payment_method, invoice.payment_method) if invoice.payment_method else ""
                    # 入帳日
                    original_ws.cell(row=row, column=5).value = invoice.account_entry_date.strftime('%Y/%m/%d') if invoice.account_entry_date else ""
                    # 金額
                    original_ws.cell(row=row, column=6).value = pi.amount
                    original_ws.cell(row=row, column=6).number_format = "#,##0"
                    # 存放行庫
                    original_ws.cell(row=row, column=7).value = bank_name
                    # 備註
                    original_ws.cell(row=row, column=8).value = invoice.notes or ""
                    # 格式設定
                    for col in range(1, 9):
                        cell = original_ws.cell(row=row, column=col)
                        cell.font = Font(size=12, name="Microsoft JhengHei", color="000000")
                        cell.alignment = Alignment(vertical="center")
                        if col in [1, 2, 3, 4, 5]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        if col == 1:
                            cell.font = Font(size=9, name="Microsoft JhengHei", color="000000")
                        cell.border = Border(bottom=Side(style="thin"))
                    original_ws.row_dimensions[row].height = 30
            else:
                # 沒有發票：只寫案號，其餘留空
                row += 1
                original_ws.cell(row=row, column=1).value = ""
                original_ws.cell(row=row, column=2).value = ""
                original_ws.cell(row=row, column=3).value = pcode
                original_ws.cell(row=row, column=4).value = ""
                original_ws.cell(row=row, column=5).value = ""
                original_ws.cell(row=row, column=6).value = ""
                original_ws.cell(row=row, column=7).value = ""
                original_ws.cell(row=row, column=8).value = ""
                # 格式設定
                for col in range(1, 9):
                    cell = original_ws.cell(row=row, column=col)
                    cell.font = Font(size=12, name="Microsoft JhengHei", color="000000")
                    cell.alignment = Alignment(vertical="center")
                    if col in [1, 2, 3, 4, 5]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(bottom=Side(style="thin"))
                original_ws.row_dimensions[row].height = 30

        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"產生請款單 Excel 發生錯誤: {str(e)}\n{traceback.format_exc()}")
        raise


def _fill_project_details(ws, project_details, start_row, custom_field_columns):
    for i, detail in enumerate(project_details):
        row = start_row + i
        item_cell = ws.cell(row=row, column=1)
        item_cell.value = detail["項次"]
        detail_cell = ws.cell(row=row, column=2)
        detail_cell.value = detail["工程明細"]
        detail_cell.alignment = Alignment(wrap_text=True, vertical="center")
        amount_cell = ws.cell(row=row, column=3)
        amount_cell.value = detail["金額"]
        amount_cell.number_format = "#,##0"
        # --- 案號格式：西元年4碼+類別代碼+3碼案號（補零） ---
        pcode_cell = ws.cell(row=row, column=4)
        project = detail["project"]
        if project:
            year = getattr(project, "year", None)
            category = getattr(project, "category", None)
            code = getattr(category, "code", "") if category else ""
            number = getattr(project, "project_number", "")
            if year and code and number is not None:
                pcode = f"{year}{code}{str(number).zfill(3)}"
                pcode_cell.value = pcode
            else:
                pcode_cell.value = ""
        # --- 自訂欄位 ---
        last_custom_col = max(custom_field_columns.values()) if custom_field_columns else None
        if project and hasattr(project, "custom_fields") and custom_field_columns:
            for field_name, col in custom_field_columns.items():
                if field_name == "pcode":
                    continue  # 已處理
                val = getattr(project, "custom_fields", {}).get(field_name)
                cell = ws.cell(row=row, column=col)
                # bool/None/空值顯示為是/否
                if project.category and project.category.custom_field_schema:
                    field_type = project.category.custom_field_schema.get(
                        field_name, {}
                    ).get("type")
                    if field_type == "boolean":
                        if val in [True, "True", "true", 1, "1"]:
                            cell.value = "是"
                        else:
                            cell.value = "否"
                    else:
                        cell.value = val if val not in [None, ""] else "否" if field_type == "boolean" else ""
                else:
                    cell.value = val if val is not None else ""
                cell.border = Border(
                    top=Side(style="mediumDashDot"),
                    bottom=Side(style="mediumDashDot"),
                )
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].hidden = False
                if project.category and project.category.custom_field_schema:
                    field_type = project.category.custom_field_schema.get(
                        field_name, {}
                    ).get("type")
                    if field_type == "number":
                        cell.number_format = "#,##0.00"
                    elif field_type == "date":
                        cell.number_format = "yyyy-mm-dd"
            # 最右邊自訂欄位的右一格填入金額
            # if last_custom_col:
            #     right_cell = ws.cell(row=row, column=last_custom_col + 1)
            #     # E7、E8欄為面積計算後得出金額，與C7、C8欄不同，應留空白
            #     # right_cell.value = detail["金額"] 
            #     right_cell.value = ""
            #     right_cell.number_format = "#,##0"


def auto_adjust_column_width(ws, column_index=None):
    for col_idx in range(1, ws.max_column + 1) if column_index is None else [column_index]:
        max_length = 0
        column = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                try:
                    cell_length = 0
                    for char in str(cell.value):
                        if ord(char) > 127:
                            cell_length += 2.1
                        else:
                            cell_length += 1.4
                    if cell.font and cell.font.bold:
                        cell_length *= 1.1
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        adjusted_width = max_length + 4
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 10 else 10

def auto_adjust_row_height(ws, row_index=None):
    for row_idx in range(1, ws.max_row + 1) if row_index is None else [row_index]:
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                line_count = str(cell.value).count('\n') + 1
                if cell.alignment and cell.alignment.wrap_text:
                    col_letter = get_column_letter(col_idx)
                    col_width = ws.column_dimensions[col_letter].width
                    if col_width:
                        chars_per_line = int(col_width / 1.2)
                        if chars_per_line > 0:
                            text_length = len(str(cell.value).replace('\n', ''))
                            estimated_lines = text_length / chars_per_line
                            line_count = max(line_count, int(estimated_lines) + 1)
                if line_count > max_lines:
                    max_lines = line_count
        row_height = max_lines * 20
        ws.row_dimensions[row_idx].height = row_height

def copy_column_and_insert(ws, source_col: int, target_col: int):
    ws.insert_cols(target_col)
    for row in range(1, ws.max_row + 1):
        src_cell = ws.cell(row=row, column=source_col)
        tgt_cell = ws.cell(row=row, column=target_col)
        tgt_cell.value = src_cell.value
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)
