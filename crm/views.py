import os
import csv
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Count, Sum, F, Q, Min, Max
from django.utils import timezone
from django.contrib import messages
from rest_framework import viewsets, filters, permissions, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.authentication import SessionAuthentication
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from .models import (
    Owner,
    Project,
    Quotation,
    Payment,
    Invoice,
    Category,
    Expenditure,
    ProjectChange,
    PaymentProject,
    Company,
    BankAccount,
    PaymentDocument,
    ProjectInvoice, 
)
from .serializers import (
    OwnerSerializer,
    ProjectSerializer,
    QuotationSerializer,
    PaymentSerializer,
    InvoiceSerializer,
    CategorySerializer,
    ExpenditureSerializer,
    ProjectChangeSerializer,
    PaymentProjectSerializer,
    CompanySerializer,
    BankAccountSerializer,
    PaymentDocumentSerializer,
    ProjectInvoiceSerializer,
)
import pandas as pd
import traceback
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from .permissions import IsAdminOrCanRequestPayment, IsAdmin
from .utils.import_utils import import_owners_from_file, import_projects_from_file
from .utils.payment_excel_service import generate_payment_excel
import logging

logger = logging.getLogger(__name__)

@login_required(login_url="signin")
def category(request):
    if not request.user.profile.is_admin:
        raise PermissionDenied
    return render(request, "crm/pages/category.html")


@login_required(login_url="signin")
def owners(request):
    return render(request, "crm/pages/owner/owners.html")


@login_required(login_url="signin")
def projects(request):
    return render(request, "crm/pages/projects.html")


@login_required(login_url="signin")
def quotations(request):
    if request.user.profile.is_admin or request.user.profile.can_request_payment:
        return render(request, "crm/pages/quotations.html")
    else:
        raise PermissionDenied
    


@login_required(login_url="signin")
def invoices(request):
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    return render(request, "crm/pages/invoices.html")


@login_required(login_url="signin")
def export(request):
    """匯出檔案頁面"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    return render(request, "crm/pages/export.html")


@login_required(login_url="signin")
def import_data(request):
    """匯入檔案頁面"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    return render(request, "crm/pages/import.html")


@login_required(login_url="signin")
def owner_projects(request, owner_id):
    """顯示特定業主的專案列表"""
    owner = get_object_or_404(Owner, id=owner_id)
    return render(request, "crm/pages/owner/owner_projects.html", {"owner": owner})


@login_required(login_url="signin")
def project_quotations(request, project_id):
    """顯示特定專案的報價列表"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    project = get_object_or_404(Project, id=project_id)
    return render(request, "crm/pages/project_quotations.html", {"project": project})


@login_required(login_url="signin")
def project_invoices(request, project_id):
    """顯示特定專案的請款列表"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    project = get_object_or_404(Project, id=project_id)
    return render(request, "crm/pages/project_invoices.html", {"project": project})


@login_required(login_url="signin")
def project_payments(request, project_id):
    """顯示特定專案的請款列表"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    project = get_object_or_404(Project, id=project_id)
    return render(request, "crm/pages/project_payments.html", {"project": project})


@login_required(login_url="signin")
def project_dashboard(request):
    """顯示專案儀表板頁面"""
    return render(request, "crm/pages/project_dashboard.html")


@login_required(login_url="signin")
def project_details(request, project_id):
    """
    顯示專案詳情頁面
    """
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        messages.error(request, "專案不存在")
        return redirect("projects")

    # 檢查用戶權限
    # 這裡可以添加額外的權限檢查，例如檢查該用戶是否有權限查看特定專案

    context = {
        "project_id": project_id,
        "page_title": f"專案詳情 - {project.name}",
    }

    return render(request, "crm/pages/project_detail.html", context)


# 請款頁面


@login_required(login_url="signin")
def create_payment(request):
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    return render(request, "crm/pages/payments/create_payment.html")


@login_required(login_url="signin")
def payments(request):
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    return render(request, "crm/pages/payments/payments.html")


@login_required(login_url="signin")
def payment_details(request, payment_id):
    """
    顯示付款單詳情頁面
    """
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    
    try:
        payment = Payment.objects.get(id=payment_id)
    except Payment.DoesNotExist:
        messages.error(request, "付款單不存在")
        return redirect("payments")

    context = {
        "payment_id": payment_id,
        "page_title": f"請款單詳情 - {payment.payment_number}",
    }

    return render(request, "crm/pages/payments/payment_detail.html", context)


@login_required(login_url="signin")
def companys(request):
    """顯示收款公司列表"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    return render(request, "crm/pages/company/companys.html")


@login_required(login_url="signin")
def company_details(request, company_id):
    """
    顯示公司詳情頁面
    """
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    
    try:
        company = Company.objects.get(id=company_id)
    except Company.DoesNotExist:
        messages.error(request, "公司不存在")
        return redirect("companys")

    context = {
        "company_id": company_id,
        "page_title": f"公司詳情 - {company.name}",
    }

    return render(request, "crm/pages/company/company_detail.html", context)


# API
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class BaseViewSet(viewsets.ModelViewSet):
    """
    基礎視圖集，提供通用的認證和權限設置
    所有 API 視圖都需要繼承此類
    """

    authentication_classes = [SessionAuthentication]
    permission_classes = [permissions.IsAuthenticated]

class AdminOnlyViewSet(viewsets.ModelViewSet):
    """
    僅限管理員使用的混合類
    """
    permission_classes = [permissions.IsAuthenticated, IsAdmin]
    
class CanPaymentViewSet(viewsets.ModelViewSet):
    """
    僅限請款人員的混合類
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOrCanRequestPayment]

class CategoryViewSet(BaseViewSet):
    queryset = Category.objects.all().order_by("code")
    serializer_class = CategorySerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["code", "description"]

    def get_permissions(self):
        """
        根據操作類型設定不同的權限。
        - 新增、更新、刪除、更新自訂欄位: 需為管理員。
        - 列表、檢視、取得自訂欄位: 需已認證。
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'update_custom_fields']:
            return [permissions.IsAuthenticated(), IsAdmin()]
        elif self.action in ['list', 'retrieve', 'custom_fields']:
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        return Category.objects.annotate(projects_count=Count("project")).order_by("code")

    @action(detail=True, methods=["get"])
    def custom_fields(self, request, pk=None):
        """取得類別的自定義欄位結構"""
        category = self.get_object()
        return Response(category.custom_field_schema or {})

    @action(detail=True, methods=["post"])
    def update_custom_fields(self, request, pk=None):
        """更新類別的自定義欄位結構"""
        category = self.get_object()
        custom_fields = request.data

        # 基本驗證
        if not isinstance(custom_fields, dict):
            return Response(
                {"error": "自定義欄位必須是物件格式"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 驗證每個欄位的格式
        for field_name, field_config in custom_fields.items():
            if not isinstance(field_config, dict):
                return Response(
                    {"error": f"欄位 {field_name} 配置必須是物件格式"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            required_keys = ["display_name", "type"]
            for key in required_keys:
                if key not in field_config:
                    return Response(
                        {"error": f"欄位 {field_name} 缺少必要的配置 {key}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            valid_types = ["text", "textarea", "number", "date", "boolean"]
            if field_config["type"] not in valid_types:
                return Response(
                    {"error": f"欄位 {field_name} 的類型不是有效類型: {valid_types}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # 更新自定義欄位結構
        category.custom_field_schema = custom_fields
        category.save(update_fields=["custom_field_schema"])

        return Response(category.custom_field_schema)


from django.db.models import ProtectedError
class OwnerViewSet(BaseViewSet):
    queryset = Owner.objects.all().order_by("tax_id")
    serializer_class = OwnerSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["company_name", "tax_id", "contact_person", "phone", "email"]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ProtectedError:
            return Response(
                {"error": "此業主尚有關聯的專案，無法刪除。"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=['get'])
    def batch_info(self, request):
        """根據ID列表批量獲取業主資訊"""
        ids = request.query_params.get('ids', '')
        if not ids:
            return Response({'error': '請提供 ids 參數'}, status=400)
        
        try:
            id_list = [int(id.strip()) for id in ids.split(',') if id.strip().isdigit()]
            if not id_list:
                return Response({'error': '無效的 ID 格式'}, status=400)
                
            owners = Owner.objects.filter(id__in=id_list).only('id', 'company_name')
            result = {owner.id: owner.company_name for owner in owners}
            return Response(result)
        except ValueError:
            return Response({'error': 'ID 必須是數字'}, status=400)


class ProjectViewSet(BaseViewSet):
    queryset = Project.objects.select_related("owner", "category").prefetch_related(
        "changes", "expenditures", "managers"
    )
    serializer_class = ProjectSerializer
    pagination_class = StandardResultsSetPagination
    # 移除 SearchFilter，因為我們在 get_queryset() 中自訂搜尋邏輯
    # filter_backends = [filters.SearchFilter]
    # search_fields = [
    #     "name",
    #     "project_number",
    #     "owner__company_name",
    #     "managers__username",
    #     "managers__profile__name",
    # ]

    def get_queryset(self):
        queryset = Project.objects.select_related("owner", "category").prefetch_related(
            "changes", "expenditures", "managers"
        )

        # 搜尋
        search_query = self.request.query_params.get("search", None)
        if search_query:
            # 基本搜尋條件
            search_conditions = (
                Q(name__icontains=search_query)
                | Q(project_number__icontains=search_query)
                | Q(owner__company_name__icontains=search_query)
                | Q(managers__username__icontains=search_query)
                | Q(managers__profile__name__icontains=search_query)
            )
            
            # 檢查是否是完整案件編號格式搜尋 (如: 2025A025)
            # 如果搜尋查詢符合年份+字母+數字的格式，嘗試拆解並搜尋
            import re
            full_number_match = re.match(r'^(\d{4})([a-zA-Z]+)(\d{3,})$', search_query)
            if full_number_match:
                year = int(full_number_match.group(1))
                category_code = full_number_match.group(2).upper()  # 統一轉為大寫比對
                project_num = full_number_match.group(3).zfill(3)  # 補齊為3位數
                
                # 添加組合搜尋條件 - 使用 |= 來添加而不是替換
                search_conditions |= Q(
                    year=year,
                    category__code=category_code,
                    project_number__icontains=project_num
                )

            queryset = queryset.filter(search_conditions).distinct()

        # 專案負責人過濾
        manager_id = self.request.query_params.get("manager", None)
        if manager_id:
            queryset = queryset.filter(managers__id=manager_id)

        # 業主過濾
        owner_id = self.request.query_params.get("owner", None)
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)

        # 類別過濾
        category_id = self.request.query_params.get("category", None)
        if category_id:
            queryset = queryset.filter(category_id=category_id)

        # 完成狀態過濾
        is_completed = self.request.query_params.get("is_completed", None)
        if is_completed is not None:
            is_completed_bool = is_completed.lower() == "true"
            queryset = queryset.filter(is_completed=is_completed_bool)

        # 發票狀態過濾
        is_invoiced = self.request.query_params.get("is_invoiced", None)
        if is_invoiced is not None:
            is_invoiced = is_invoiced.lower() == "true"
            queryset = queryset.filter(is_invoiced=is_invoiced)

        # 付款狀態過濾
        is_paid = self.request.query_params.get("is_paid", None)
        if is_paid is not None:
            is_paid = is_paid.lower() == "true"
            queryset = queryset.filter(is_paid=is_paid)

        # 年份過濾 - 單一年份
        year = self.request.query_params.get("year", None)
        if year:
            queryset = queryset.filter(year=year)

        # 年份區間過濾 - 開始年份
        year_start = self.request.query_params.get("year_start", None)
        if year_start:
            queryset = queryset.filter(year__gte=year_start)

        # 年份區間過濾 - 結束年份
        year_end = self.request.query_params.get("year_end", None)
        if year_end:
            queryset = queryset.filter(year__lte=year_end)
        
        return queryset.order_by("-year", "project_number")

    @action(detail=False, methods=["get"])
    def years(self, request):
        """獲取所有可用的專案年份"""
        years = (
            Project.objects.values_list("year", flat=True).distinct().order_by("-year")
        )

        # 獲取最小和最大年份
        min_year = Project.objects.aggregate(Min("year"))["year__min"]
        max_year = Project.objects.aggregate(Max("year"))["year__max"]

        # 如果沒有專案，設定預設值
        if not min_year:
            current_year = timezone.now().year
            min_year = current_year
            max_year = current_year

        return Response(
            {"years": list(years), "min_year": min_year, "max_year": max_year}
        )

    @action(detail=False, methods=["get"])
    def dashboard(self, request):
        """獲取專案儀表板所需的統計數據"""
        # 獲取基本統計數據
        total_projects = Project.objects.count()
        active_projects = Project.objects.filter(is_completed=False).count()
        completed_projects = Project.objects.filter(is_completed=True).count()

        # 計算百分比
        completion_rate = (
            round((completed_projects / total_projects) * 100)
            if total_projects > 0
            else 0
        )
        active_projects_percent = (
            round((active_projects / total_projects) * 100) if total_projects > 0 else 0
        )
        completed_projects_percent = (
            round((completed_projects / total_projects) * 100)
            if total_projects > 0
            else 0
        )

        # 營收數據
        total_quotations = Quotation.objects.all()
        total_revenue = Quotation.objects.aggregate(total=Sum("amount"))["total"] or 0
        paid_revenue = (
            Invoice.objects.filter(paid=True).aggregate(total=Sum("amount"))["total"]
            or 0
        )
        unpaid_revenue = total_revenue - paid_revenue

        # 類別分布數據
        category_data = []
        categories = Category.objects.all().order_by("code")
        for category in categories:
            category_count = Project.objects.filter(category=category).count()
            if category_count > 0:
                category_data.append(
                    {
                        "name": f"{category.code}: {category.description}",
                        "count": category_count,
                    }
                )

        # 年度趨勢數據
        yearly_data = []
        years = (
            Project.objects.values("year").annotate(count=Count("id")).order_by("year")
        )
        for year_data in years:
            yearly_data.append({"year": year_data["year"], "count": year_data["count"]})

        # 整合數據
        stats = {
            "totalProjects": total_projects,
            "activeProjects": active_projects,
            "completedProjects": completed_projects,
            "completionRate": completion_rate,
            "activeProjectsPercent": active_projects_percent,
            "completedProjectsPercent": completed_projects_percent,
            "totalRevenue": float(total_revenue),
            "paidRevenue": float(paid_revenue),
            "unpaidRevenue": float(unpaid_revenue),
        }

        return Response(
            {
                "stats": stats,
                "categoryData": category_data,
                "yearlyData": yearly_data,
            }
        )

    def perform_create(self, serializer):
        """執行創建操作，處理 Model 層的 ValidationError"""
        from django.core.exceptions import ValidationError as DjangoValidationError
        
        try:
            serializer.save()
        except DjangoValidationError as e:
            # 重新拋出為 DRF 的 ValidationError
            from rest_framework.exceptions import ValidationError
            raise ValidationError({"error": str(e)})

    def perform_update(self, serializer):
        """執行更新操作，處理 Model 層的 ValidationError"""
        from django.core.exceptions import ValidationError as DjangoValidationError
        
        try:
            serializer.save()
        except DjangoValidationError as e:
            # 重新拋出為 DRF 的 ValidationError
            from rest_framework.exceptions import ValidationError
            raise ValidationError({"error": str(e)})

    def create(self, request, *args, **kwargs):
        """創建專案，包含案件編號重複檢查"""
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        """更新專案，包含案件編號重複檢查"""
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # 確保返回更新後的完整資料，包括 managers_info
        return Response(self.get_serializer(instance).data)


class QuotationViewSet(BaseViewSet):
    queryset = Quotation.objects.all().select_related("project")
    serializer_class = QuotationSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]

    def get_queryset(self):
        queryset = Quotation.objects.all().select_related("project", "project__owner")

        # 專案過濾
        project_id = self.request.query_params.get("project", None)
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        return queryset.order_by("-date_issued")


class PaymentViewSet(CanPaymentViewSet):
    queryset = Payment.objects.all().prefetch_related(
        "projects", "paymentproject_set", "paymentproject_set__project", 
        "paymentproject_set__project__owner"
    )
    serializer_class = PaymentSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["payment_number", "projects__name", "projects__owner__company_name"]

    def get_queryset(self):
        queryset = Payment.objects.all().prefetch_related(
            "projects", "paymentproject_set", "paymentproject_set__project",
            "paymentproject_set__project__owner"
        )

        # 報價過濾
        quotation_id = self.request.query_params.get("quotation", None)
        if quotation_id:
            queryset = queryset.filter(paymentproject__quotation_id=quotation_id)

        # 專案過濾
        project_id = self.request.query_params.get("project", None)
        if project_id:
            queryset = queryset.filter(projects__id=project_id)

        # 客戶過濾
        owner_id = self.request.query_params.get("owner", None)
        if owner_id:
            queryset = queryset.filter(projects__owner_id=owner_id)
        
        # 付款狀態過濾
        paid = self.request.query_params.get("paid", None)
        if paid is not None:
            if paid.lower() == "true":
                queryset = queryset.filter(paid=True)
            elif paid.lower() == "false":
                queryset = queryset.filter(paid=False)

        return queryset.order_by("-date_issued")

    def create(self, request, *args, **kwargs):
        """
        建立新請款單，同時處理關聯的專案
        """
        # 建立請款單
        payment_serializer = self.get_serializer(data=request.data)
        payment_serializer.is_valid(raise_exception=True)
        payment = payment_serializer.save(created_by=request.user, amount=0)

        # 處理關聯的專案
        payment_projects = request.data.get("payment_projects", [])
        total_amount = 0

        for project_item in payment_projects:
            project_data = {
                "payment": payment.id,
                "project": project_item.get("project"),
                "quotation": project_item.get("quotation", None),
                "amount": project_item.get("amount"),
                "description": project_item.get("description", ""),
            }

            # 建立請款單-專案關聯
            payment_project_serializer = PaymentProjectSerializer(data=project_data)
            if payment_project_serializer.is_valid():
                payment_project_serializer.save()
                total_amount += float(project_item.get("amount", 0))
            else:
                payment.delete()
                return Response(
                    payment_project_serializer.errors,
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # 更新請款單總金額
        payment.amount = total_amount
        payment.save()

        return Response(
            self.get_serializer(payment).data, status=status.HTTP_201_CREATED
        )

    def partial_update(self, request, *args, **kwargs):
        """支援 PATCH 方法來更新付款狀態"""
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class PaymentProjectViewSet(CanPaymentViewSet):
    queryset = PaymentProject.objects.all().select_related("project", "payment")
    serializer_class = PaymentProjectSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = PaymentProject.objects.all().select_related("project", "payment")

        # 請款單過濾
        payment_id = self.request.query_params.get("payment", None)
        if payment_id:
            queryset = queryset.filter(payment_id=payment_id)

        # 專案過濾
        project_id = self.request.query_params.get("project", None)
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # 預加載專案變更次數
        queryset = queryset.prefetch_related("project__changes")

        return queryset

    def perform_create(self, serializer):
        """創建時自動更新請款單總金額"""
        payment_project = serializer.save()
        payment_project.payment.update_amount()

    def perform_update(self, serializer):
        """更新時自動更新請款單總金額"""
        payment_project = serializer.save()
        payment_project.payment.update_amount()

    def perform_destroy(self, instance):
        """刪除時自動更新請款單總金額"""
        payment = instance.payment
        instance.delete()
        payment.update_amount()


class InvoiceViewSet(CanPaymentViewSet):
    queryset = Invoice.objects.all().select_related("payment", "created_by")
    serializer_class = InvoiceSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["invoice_number", "payment__payment_number"]

    def get_queryset(self):
        queryset = Invoice.objects.all().select_related("payment", "created_by")

        # 請款單過濾
        payment_id = self.request.query_params.get("payment", None)
        if payment_id:
            queryset = queryset.filter(payment_id=payment_id)
            
        # 付款狀態過濾（新版使用 payment_status，舊版使用 is_paid）
        payment_status = self.request.query_params.get("payment_status", None)
        is_paid = self.request.query_params.get("is_paid", None)
        
        if payment_status:
            # 使用新的 payment_status 欄位進行篩選
            queryset = queryset.filter(payment_status=payment_status)
        elif is_paid is not None:
            # 向後相容：支援舊的 is_paid 參數
            is_paid_bool = is_paid.lower() == 'true'
            if is_paid_bool:
                queryset = queryset.filter(payment_status='paid')
            else:
                queryset = queryset.exclude(payment_status='paid')
            
        # 付款方式過濾
        payment_method = self.request.query_params.get("payment_method", None)
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)
            
        # 發票類型過濾
        invoice_type = self.request.query_params.get("invoice_type", None)
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)
            
        # 日期範圍過濾
        issue_date_gte = self.request.query_params.get("issue_date__gte", None)
        if issue_date_gte:
            queryset = queryset.filter(issue_date__gte=issue_date_gte)
            
        issue_date_lte = self.request.query_params.get("issue_date__lte", None)
        if issue_date_lte:
            queryset = queryset.filter(issue_date__lte=issue_date_lte)

        # 排序參數
        ordering = self.request.query_params.get('ordering', '-issue_date')
        return queryset.order_by(ordering)

    def perform_create(self, serializer):
        # 先建立發票
        invoice = serializer.save(created_by=self.request.user)
        # 取得前端送來的 project_amounts
        project_amounts = self.request.data.get("project_amounts", [])
        for item in project_amounts:
            project_id = item.get("project_id")
            amount = item.get("amount")
            if project_id:
                # 建立 ProjectInvoice 關聯
                ProjectInvoice.objects.create(
                    invoice=invoice,
                    project_id=project_id,
                    amount=amount
                )

    def perform_update(self, serializer):
        # 更新發票
        invoice = serializer.save()
        # 取得前端送來的 project_amounts
        project_amounts = self.request.data.get("project_amounts", [])
        
        # 先刪除舊的關聯
        ProjectInvoice.objects.filter(invoice=invoice).delete()
        
        # 建立新的關聯
        for item in project_amounts:
            project_id = item.get("project_id")
            amount = item.get("amount")
            if project_id:
                ProjectInvoice.objects.create(
                    invoice=invoice,
                    project_id=project_id,
                    amount=amount
                )


# 案件支出
class ExpenditureViewSet(BaseViewSet):
    queryset = Expenditure.objects.all().select_related("project", "created_by")
    serializer_class = ExpenditureSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["description"]

    def get_queryset(self):
        queryset = Expenditure.objects.all().select_related("project", "created_by")

        # 專案過濾
        project_id = self.request.query_params.get("project", None)
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        return queryset.order_by("-date")

    def perform_create(self, serializer):
        # 自動設置建立者為當前用戶
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        # 更新時保留原始建立者
        serializer.save()


# 案件變更
class ProjectChangeViewSet(BaseViewSet):
    queryset = ProjectChange.objects.all().select_related("project", "created_by")
    serializer_class = ProjectChangeSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["description"]

    def get_queryset(self):
        queryset = ProjectChange.objects.all().select_related("project", "created_by")

        # 專案過濾
        project_id = self.request.query_params.get("project", None)
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        return queryset.order_by("-created_at")

    def perform_create(self, serializer):
        # 自動設置建立者為當前用戶
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        # 更新時保留原始建立者
        serializer.save()


class CompanyViewSet(BaseViewSet):
    queryset = Company.objects.all().order_by("tax_id")
    serializer_class = CompanySerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "tax_id", "contact_person", "phone"]

    def get_permissions(self):
        """
        company 相關操作（新增、查詢、修改、刪除、銀行帳戶）允許 admin 或 can_request_payment
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'bank_accounts', 'add_bank_account']:
            return [permissions.IsAuthenticated(), IsAdminOrCanRequestPayment()]
        return [permissions.IsAuthenticated()]
        
    def get_queryset(self):
        return Company.objects.all().prefetch_related('bank_accounts', 'payments').order_by("tax_id")
        
    @action(detail=True, methods=['get'])
    def bank_accounts(self, request, pk=None):
        """獲取公司的銀行帳戶"""
        company = self.get_object()
        bank_accounts = company.bank_accounts.all()
        serializer = BankAccountSerializer(bank_accounts, many=True)
        return Response(serializer.data)
        
    @action(detail=True, methods=['post'])
    def add_bank_account(self, request, pk=None):
        """新增銀行帳戶到公司"""
        company = self.get_object()
        
        # 將公司ID添加到請求數據中
        request_data = request.data.copy()
        request_data['company'] = company.id
        
        serializer = BankAccountSerializer(data=request_data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BankAccountViewSet(BaseViewSet):
    queryset = BankAccount.objects.all().select_related('company')
    serializer_class = BankAccountSerializer
    pagination_class = StandardResultsSetPagination
    
    def get_permissions(self):
        """
        確保只有管理員可以訪問銀行帳戶資訊
        """
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdmin()]
        return [permissions.IsAuthenticated()]
        
    def get_queryset(self):
        queryset = BankAccount.objects.all().select_related('company')
        
        # 根據公司過濾
        company_id = self.request.query_params.get('company', None)
        if company_id:
            queryset = queryset.filter(company_id=company_id)
            
        return queryset


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminOrCanRequestPayment])
def import_owners(request):
    """匯入業主資料"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        return Response({"detail": "您沒有權限執行此操作"}, status=403)

    try:
        # 確保請求中包含檔案
        if 'file' not in request.FILES:
            return Response({"detail": "請求中未包含檔案"}, status=400)

        file = request.FILES['file']

        # 呼叫匯入工具函式
        result = import_owners_from_file(file)

        return Response(result, status=200)
    except Exception as e:
        return Response({"detail": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminOrCanRequestPayment])
def import_projects(request):
    """匯入專案資料"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        return Response({"detail": "您沒有權限執行此操作"}, status=403)

    try:
        # 確保請求中包含檔案
        if 'file' not in request.FILES:
            return Response({"detail": "請求中未包含檔案"}, status=400)

        file = request.FILES['file']

        # 呼叫匯入工具函式
        result = import_projects_from_file(file)

        return Response(result, status=200)
    except Exception as e:
        return Response({"detail": str(e)}, status=500)


from rest_framework.decorators import api_view, permission_classes

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsAdminOrCanRequestPayment])
def get_bank_accounts_for_company(request, company_id):
    """根據公司ID獲取該公司的所有銀行帳號"""
    try:
        company = get_object_or_404(Company, id=company_id)
        bank_accounts = BankAccount.objects.filter(company=company)
        serializer = BankAccountSerializer(bank_accounts, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@login_required(login_url="signin")
def index(request):
    return render(request, "crm/index.html")


@login_required(login_url="signin")
def export_payment_excel(request, payment_id):
    """匯出請款單為Excel檔案"""
    try:
        payment = get_object_or_404(Payment, id=payment_id)
        output = generate_payment_excel(payment)
        filename = f"payment_{payment.payment_number}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except FileNotFoundError as e:
        logger.error(f"匯出Excel失敗: {str(e)}")
        return HttpResponse(f"匯出Excel失敗：找不到樣板檔案，請聯絡管理員。", status=500)
    except ValueError as e:
        logger.error(f"匯出Excel失敗: {str(e)}")
        return HttpResponse(f"匯出Excel失敗：資料不完整，{str(e)}", status=500)
    except Exception as e:
        import traceback
        logger.error(f"匯出Excel失敗: {str(e)}\n{traceback.format_exc()}")
        return HttpResponse(f"匯出Excel失敗: {str(e)}", status=500)
        
def copy_column_and_insert(ws, source_col: int, target_col: int):
    """
    將指定欄位（source_col）複製到指定欄位位置（target_col），會自動先插入欄位
    :param ws: openpyxl 的工作表物件
    :param source_col: 要複製的來源欄位編號（從1開始）
    :param target_col: 要插入並貼上的目標欄位編號（從1開始）
    """
    # Step 1: 插入空白欄，位置就是 target_col
    ws.insert_cols(target_col)

    # Step 2: 開始逐列複製資料與格式
    for row in range(1, ws.max_row + 1):
        src_cell = ws.cell(row=row, column=source_col)
        tgt_cell = ws.cell(row=row, column=target_col)

        tgt_cell.value = src_cell.value
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)

@login_required(login_url="signin")
def export_projects_csv(request):
    """匯出所有專案資料為 CSV 格式，支援年份篩選"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    import csv
    from datetime import datetime
    now = timezone.now()
    filename = f"專案資料_{now.strftime('%Y%m%d-%H%M%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    from urllib.parse import quote
    ascii_filename = f"projects_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
    response.write('\ufeff')
    writer = csv.writer(response)
    
    # 取得查詢參數（只保留年份篩選）
    year_start = request.GET.get('year_start')
    year_end = request.GET.get('year_end')
    
    projects = Project.objects.all().select_related('owner', 'category').prefetch_related('managers', 'quotations')
    
    # 年份篩選
    if year_start:
        projects = projects.filter(year__gte=year_start)
    if year_end:
        projects = projects.filter(year__lte=year_end)
    headers = [
        '年份', '案件類別', '案件編號', '案件名稱', '業主', 
        '負責人', '繪圖', '聯絡方式', '報價', '是否完成', 
        '請款日期', '請款金額', '收款日期', '發票日期', 
        '是否請款', '是否收款', '備註'
    ]
    writer.writerow(headers)
    for project in projects:
        managers_names = ', '.join([
            manager.profile.name if hasattr(manager, 'profile') and manager.profile.name else manager.username
            for manager in project.managers.all()
        ])
        quotations_list = project.quotations.all()
        quotations_text = ', '.join([f'${{q.amount}}' for q in quotations_list]) if quotations_list.exists() else ''
        row = [
            project.year,
            f"{project.category.code}:{project.category.description}" if project.category else '',
            str(project.project_number) if project.project_number else '',
            project.name,
            project.owner.company_name if project.owner else '',
            managers_names,
            project.drawing,
            project.contact_info,
            quotations_text,
            '是' if project.is_completed else '否',
            project.invoice_date.strftime('%Y-%m-%d') if project.invoice_date else '',
            project.invoice_amount if project.invoice_amount else '',
            project.payment_date.strftime('%Y-%m-%d') if project.payment_date else '',
            project.invoice_issue_date.strftime('%Y-%m-%d') if project.invoice_issue_date else '',
            '是' if project.is_invoiced else '否',
            '是' if project.is_paid else '否',
            project.notes
        ]
        writer.writerow(row)
    return response

@login_required(login_url="signin")
def export_owners_csv(request):
    """匯出所有業主資料為 CSV 格式（不再依年份過濾）"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    import csv
    now = timezone.now()
    filename = f"業主資料_{now.strftime('%Y%m%d-%H%M%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    from urllib.parse import quote
    ascii_filename = f"owners_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
    response.write('\ufeff')
    writer = csv.writer(response)
    owners = Owner.objects.all()
    headers = [
        '公司名稱', '統一編號', '電話', '傳真', 
        '電子信箱', '手機', '地址', '聯絡人'
    ]
    writer.writerow(headers)
    for owner in owners:
        row = [
            owner.company_name,
            owner.tax_id,
            owner.phone,
            owner.fax if owner.fax else '',
            owner.email,
            owner.mobile if owner.mobile else '',
            owner.address,
            owner.contact_person
        ]
        writer.writerow(row)
    return response

@login_required(login_url="signin")
def export_quotations_csv(request):
    """匯出所有報價資料為 CSV 格式，支援年份篩選（依專案年份）"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    import csv
    now = timezone.now()
    filename = f"報價資料_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    from urllib.parse import quote
    ascii_filename = f"quotations_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
    response.write('\ufeff')
    writer = csv.writer(response)
    year_start = request.GET.get('year_start')
    year_end = request.GET.get('year_end')
    quotations = Quotation.objects.all().select_related('project', 'project__owner', 'project__category')
    if year_start:
        quotations = quotations.filter(project__year__gte=year_start)
    if year_end:
        quotations = quotations.filter(project__year__lte=year_end)
    headers = [
        '專案年份', '專案類別', '專案編號', '專案名稱', '業主', 
        '報價金額', '報價發行日期'
    ]
    writer.writerow(headers)
    for quotation in quotations:
        project = quotation.project
        row = [
            project.year,
            f"{project.category.code}:{project.category.description}" if project.category else '',
            str(project.project_number) if project.project_number else '',
            project.name,
            project.owner.company_name if project.owner else '',
            quotation.amount,
            quotation.date_issued.strftime('%Y-%m-%d') if quotation.date_issued else ''
        ]
        writer.writerow(row)
    return response

@login_required(login_url="signin")
def export_payments_csv(request):
    """匯出所有請款資料為 CSV 格式，支援年份篩選（依專案年份）和年月篩選（依請款日期）"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    import csv
    from datetime import datetime
    now = timezone.now()
    filename = f"請款資料_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    from urllib.parse import quote
    ascii_filename = f"payments_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
    response.write('\ufeff')
    writer = csv.writer(response)
    
    # 取得查詢參數
    year_start = request.GET.get('year_start')
    year_end = request.GET.get('year_end')
    year_month_start = request.GET.get('year_month_start')
    year_month_end = request.GET.get('year_month_end')
    
    payments = Payment.objects.all().prefetch_related('projects', 'paymentproject_set', 'paymentproject_set__project').select_related('owner', 'company', 'selected_bank_account', 'created_by')
    
    # 年份篩選（原有邏輯，依專案年份）
    if year_start:
        payments = payments.filter(projects__year__gte=year_start)
    if year_end:
        payments = payments.filter(projects__year__lte=year_end)
    
    # 年月篩選（基於請款發行日期 date_issued）
    if year_month_start:
        try:
            year, month = year_month_start.split('-')
            start_date = datetime(int(year), int(month), 1).date()
            payments = payments.filter(date_issued__gte=start_date)
        except (ValueError, TypeError):
            pass  # 忽略無效的日期格式
    
    if year_month_end:
        try:
            year, month = year_month_end.split('-')
            # 計算該月的最後一天
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1).date()
            else:
                end_date = datetime(int(year), int(month) + 1, 1).date()
            from datetime import timedelta
            end_date = end_date - timedelta(days=1)
            payments = payments.filter(date_issued__lte=end_date)
        except (ValueError, TypeError):
            pass  # 忽略無效的日期格式
            
    payments = payments.distinct()
    headers = [
        '請款單號', '業主', '收款公司', '請款金額', '請款日期', 
        '到期日', '是否已付款', '付款日', '備註', '建立者', '建立時間'
    ]
    writer.writerow(headers)
    for payment in payments:
        row = [
            payment.payment_number,
            payment.owner.company_name if payment.owner else '',
            payment.company.name if payment.company else '',
            payment.amount,
            payment.date_issued.strftime('%Y-%m-%d') if payment.date_issued else '',
            payment.due_date.strftime('%Y-%m-%d') if payment.due_date else '',
            '是' if payment.paid else '否',
            payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
            payment.notes,
            payment.created_by.profile.name if payment.created_by and hasattr(payment.created_by, 'profile') and payment.created_by.profile.name else (payment.created_by.username if payment.created_by else ''),
            payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else ''
        ]
        writer.writerow(row)
    return response

@login_required(login_url="signin")
def export_invoices_csv(request):
    """匯出所有發票資料為 CSV 格式，支援年份篩選（依請款單關聯專案年份）和年月篩選（依發票開立日期）"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    import csv
    from datetime import datetime, timedelta
    now = timezone.now()
    filename = f"發票資料_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    from urllib.parse import quote
    ascii_filename = f"invoices_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
    response.write('\ufeff')
    writer = csv.writer(response)
    
    # 取得查詢參數
    year_start = request.GET.get('year_start')
    year_end = request.GET.get('year_end')
    year_month_start = request.GET.get('year_month_start')
    year_month_end = request.GET.get('year_month_end')
    
    invoices = Invoice.objects.all().select_related('payment', 'payment__owner', 'payment__company', 'created_by')
    
    # 年月篩選（基於發票開立日期 issue_date）
    if year_month_start:
        try:
            year, month = year_month_start.split('-')
            start_date = datetime(int(year), int(month), 1).date()
            invoices = invoices.filter(issue_date__gte=start_date)
        except (ValueError, TypeError):
            pass  # 忽略無效的日期格式
    
    if year_month_end:
        try:
            year, month = year_month_end.split('-')
            # 計算該月的最後一天
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1).date()
            else:
                end_date = datetime(int(year), int(month) + 1, 1).date()
            end_date = end_date - timedelta(days=1)
            invoices = invoices.filter(issue_date__lte=end_date)
        except (ValueError, TypeError):
            pass  # 忽略無效的日期格式
    
    # 年份篩選（原有邏輯，依請款單關聯專案年份）
    if year_start or year_end:
        # 只篩選有 payment 且 payment 有 projects
        filtered_invoices = []
        for invoice in invoices:
            payment = invoice.payment
            if not payment:
                continue
            projects = payment.projects.all()
            if not projects.exists():
                continue
            years = [p.year for p in projects if p.year]
            if not years:
                continue
            if year_start and all(y < int(year_start) for y in years):
                continue
            if year_end and all(y > int(year_end) for y in years):
                continue
            filtered_invoices.append(invoice)
        invoices = filtered_invoices
    else:
        # 如果沒有年份篩選，則轉換為 QuerySet 列表以保持一致性
        invoices = list(invoices)
    headers = [
        '發票號碼', '請款單號', '業主', '收款公司', '請款金額', 
        '收款日', '入帳日', '收款方式', 
        '實收金額', '付款狀態', '備註', '建立者', '建立時間'
    ]
    writer.writerow(headers)
    for invoice in invoices:
        payment_method_display = ''
        if invoice.payment_method:
            payment_method_choices = dict(Invoice.PAYMENT_METHOD_CHOICES)
            payment_method_display = payment_method_choices.get(invoice.payment_method, invoice.payment_method)
        row = [
            invoice.invoice_number,
            invoice.payment.payment_number if invoice.payment else '',
            invoice.payment.owner.company_name if invoice.payment and invoice.payment.owner else '',
            invoice.payment.company.name if invoice.payment and invoice.payment.company else '',
            invoice.amount,
            invoice.payment_received_date.strftime('%Y-%m-%d') if invoice.payment_received_date else '',
            invoice.account_entry_date.strftime('%Y-%m-%d') if invoice.account_entry_date else '',
            payment_method_display,
            invoice.actual_received_amount if invoice.actual_received_amount else '',
            invoice.get_payment_status_display() if invoice.payment_status else '',
            invoice.notes if invoice.notes else '',
            invoice.created_by.profile.name if invoice.created_by and hasattr(invoice.created_by, 'profile') and invoice.created_by.profile.name else (invoice.created_by.username if invoice.created_by else ''),
            invoice.created_at.strftime('%Y-%m-%d %H:%M:%S') if invoice.created_at else ''
        ]
        writer.writerow(row)
    return response

@login_required(login_url="signin")
def export_categories_csv(request):
    """匯出所有案件類別資料為 CSV 格式（不再依年份過濾）"""
    if not request.user.profile.is_admin and not request.user.profile.can_request_payment:
        raise PermissionDenied
    import csv
    now = timezone.now()
    filename = f"案件類別_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    from urllib.parse import quote
    ascii_filename = f"categories_{now.strftime('%Y%m%d-%H%H%S')}.csv"
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
    response.write('\ufeff')
    writer = csv.writer(response)
    categories = Category.objects.all()
    headers = [
        '類別代碼', '類別說明', '自定義欄位'
    ]
    writer.writerow(headers)
    for category in categories:
        custom_fields_display = ''
        if category.custom_field_schema:
            field_names = []
            for field_name, field_config in category.custom_field_schema.items():
                display_name = field_config.get('display_name', field_name)
                field_type = field_config.get('type', 'text')
                required = '必填' if field_config.get('required', False) else '選填'
                field_names.append(f"{display_name}({field_type},{required})")
            custom_fields_display = '; '.join(field_names)
        row = [
            category.code,
            category.description,
            custom_fields_display
        ]
        writer.writerow(row)
    return response


# 匯入功能 API endpoints
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminOrCanRequestPayment])
@parser_classes([MultiPartParser, FormParser])
def import_owners_api(request):
    """業主資料匯入 API"""
    if 'file' not in request.FILES:
        return Response(
            {'error': '請選擇要匯入的檔案'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    uploaded_file = request.FILES['file']
    
    # 檢查檔案類型
    file_extension = uploaded_file.name.lower().split('.')[-1]
    if file_extension not in ['csv', 'xlsx', 'xls']:
        return Response(
            {'error': '不支援的檔案格式，請使用 CSV 或 Excel 檔案'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # 儲存暫存檔案
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_extension}') as temp_file:
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)
            temp_file_path = temp_file.name
        
        # 執行匯入
        file_type = 'excel' if file_extension in ['xlsx', 'xls'] else 'csv'
        result = import_owners_from_file(temp_file_path, file_type)
        
        # 清理暫存檔案
        os.unlink(temp_file_path)
        
        return Response({
            'success': True,
            'message': f'匯入完成：成功 {result.success_count} 筆，錯誤 {result.error_count} 筆',
            'result': result.to_dict()
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        # 清理暫存檔案（如果存在）
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
        
        return Response(
            {'error': f'匯入過程中發生錯誤: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminOrCanRequestPayment])
@parser_classes([MultiPartParser, FormParser])
def import_projects_api(request):
    """專案資料匯入 API"""
    if 'file' not in request.FILES:
        return Response(
            {'error': '請選擇要匯入的檔案'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    uploaded_file = request.FILES['file']
    
    # 檢查檔案類型
    file_extension = uploaded_file.name.lower().split('.')[-1]
    if file_extension not in ['csv', 'xlsx', 'xls']:
        return Response(
            {'error': '不支援的檔案格式，請使用 CSV 或 Excel 檔案'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # 儲存暫存檔案
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_extension}') as temp_file:
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)
            temp_file_path = temp_file.name
        
        # 執行匯入
        file_type = 'excel' if file_extension in ['xlsx', 'xls'] else 'csv'
        result = import_projects_from_file(temp_file_path, file_type)
        
        # 清理暫存檔案
        os.unlink(temp_file_path)
        
        return Response({
            'success': True,
            'message': f'匯入完成：成功 {result.success_count} 筆，錯誤 {result.error_count} 筆',
            'result': result.to_dict()
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        # 清理暫存檔案（如果存在）
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
        
        return Response(
            {'error': f'匯入過程中發生錯誤: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class PaymentDocumentViewSet(CanPaymentViewSet):
    """內存請款單檔案視圖集"""
    
    queryset = PaymentDocument.objects.all().select_related('payment', 'uploaded_by')
    serializer_class = PaymentDocumentSerializer
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # 根據請款單 ID 過濾
        payment_id = self.request.query_params.get('payment', None)
        if payment_id:
            queryset = queryset.filter(payment_id=payment_id)
        
        return queryset.order_by('-uploaded_at')
    
    def create(self, request, *args, **kwargs):
        """上傳檔案"""
        try:
            # 取得請款單 ID
            payment_id = request.data.get('payment')
            if not payment_id:
                return Response(
                    {'error': '請指定請款單'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 驗證請款單是否存在
            try:
                payment = Payment.objects.get(id=payment_id)
            except Payment.DoesNotExist:
                return Response(
                    {'error': '請款單不存在'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # 取得上傳的檔案
            file = request.FILES.get('file')
            if not file:
                return Response(
                    {'error': '請選擇要上傳的檔案'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 檢查檔案大小（1MB 限制）
            if file.size > 1024 * 1024:
                return Response(
                    {'error': '檔案大小不能超過 1MB'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 建立 PaymentDocument 物件
            document = PaymentDocument(
                payment=payment,
                original_filename=file.name,
                uploaded_by=request.user
            )
            document.file = file
            document.save()
            
            # 序列化並回傳結果
            serializer = self.get_serializer(document)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'上傳失敗: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """刪除檔案"""
        try:
            instance = self.get_object()
            
            # 刪除實際檔案
            if instance.file:
                if os.path.exists(instance.file.path):
                    os.remove(instance.file.path)
            
            # 刪除資料庫記錄
            instance.delete()
            
            return Response(
                {'message': '檔案已成功刪除'}, 
                status=status.HTTP_200_OK
            )
            
        except Exception as e:            return Response(
                {'error': f'刪除失敗: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """下載檔案"""
        try:
            document = self.get_object()
            
            if not document.file or not os.path.exists(document.file.path):
                return Response(
                    {'error': '檔案不存在'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # 取得原始檔名
            filename = document.get_display_filename()
            
            # 處理檔名編碼 - 支援中文檔名
            from urllib.parse import quote
            
            # 建立檔案回應
            with open(document.file.path, 'rb') as f:
                response = HttpResponse(
                    f.read(), 
                    content_type='application/octet-stream'
                )
                
                # 設定檔名，支援中文 - 使用RFC 5987標準
                try:
                    # 嘗試ASCII編碼
                    filename.encode('ascii')
                    # 如果成功，直接使用檔名
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                except UnicodeEncodeError:
                    # 如果包含非ASCII字符，使用UTF-8編碼
                    encoded_filename = quote(filename.encode('utf-8'))
                    response['Content-Disposition'] = f'attachment; filename="download"; filename*=UTF-8\'\'{encoded_filename}'
                
                return response
                
        except Exception as e:
            return Response(
                {'error': f'下載失敗: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ProjectInvoiceViewSet(viewsets.ModelViewSet):
    queryset = ProjectInvoice.objects.select_related("invoice", "project")
    serializer_class = ProjectInvoiceSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = ProjectInvoice.objects.select_related("invoice", "project")
        invoice_id = self.request.query_params.get("invoice", None)
        project_id = self.request.query_params.get("project", None)
        if invoice_id:
            queryset = queryset.filter(invoice_id=invoice_id)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset.order_by("-id")


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsAdminOrCanRequestPayment])
def export_payment_invoices_excel(request):
    """
    匯出請款單發票資料為Excel格式
    根據時間區間匯出請款單的發票資料，每個專案顯示一筆記錄
    """
    try:
        # 取得篩選參數
        date_start = request.GET.get('date_start')  # 格式: YYYY-MM-DD
        date_end = request.GET.get('date_end')      # 格式: YYYY-MM-DD
        
        # 建立基礎查詢
        payment_filter = Q()
        
        # 如果有提供日期範圍，根據請款發行日期篩選
        if date_start:
            try:
                payment_filter &= Q(date_issued__gte=date_start)
            except ValueError:
                return Response({'error': '開始日期格式不正確，請使用 YYYY-MM-DD 格式'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        if date_end:
            try:
                payment_filter &= Q(date_issued__lte=date_end)
            except ValueError:
                return Response({'error': '結束日期格式不正確，請使用 YYYY-MM-DD 格式'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        # 查詢請款單及相關資料
        payments = Payment.objects.filter(payment_filter).select_related(
            'owner', 'company', 'created_by'
        ).prefetch_related(
            'paymentproject_set__project__category',
            'invoices',
            'projects__projectinvoice_set'
        ).order_by('date_issued', 'payment_number')
        
        if not payments.exists():
            return Response({'message': '在指定的時間範圍內沒有找到請款單資料'}, 
                          status=status.HTTP_204_NO_CONTENT)
        
        # 準備匯出資料
        export_data = []
        
        for payment in payments:
            # 遍歷請款單中的每個專案
            for payment_project in payment.paymentproject_set.all():
                project = payment_project.project
                
                if not project:
                    continue
                
                # 組合案件編號 (year + category_code + project_number)
                category_code = project.category.code if project.category else 'N'
                project_code = f"{project.year}{category_code}{project.project_number or '000'}"
                
                # 查找該專案對應的發票和ProjectInvoice記錄
                project_invoice_record = None
                corresponding_invoice = None
                
                # 找出該專案在這個請款單中對應的發票
                for invoice in payment.invoices.all():
                    project_invoice = invoice.projectinvoice_set.filter(project=project).first()
                    if project_invoice:
                        project_invoice_record = project_invoice
                        corresponding_invoice = invoice
                        break
                
                # 如果沒有找到對應的發票記錄，使用預設值
                if not corresponding_invoice:
                    # 預設值處理
                    payment_method_display = '-'
                    payment_status_display = '未付款'
                    project_invoice_amount = 0
                    payment_received_date_str = '-'
                    account_entry_date_str = '-'
                    notes = ''
                else:
                    # 格式化收款方式
                    if corresponding_invoice.payment_method:
                        payment_method_choices = dict(Invoice.PAYMENT_METHOD_CHOICES)
                        payment_method_display = payment_method_choices.get(
                            corresponding_invoice.payment_method, 
                            corresponding_invoice.payment_method
                        )
                    else:
                        payment_method_display = '-'
                    
                    # 格式化付款狀態
                    if corresponding_invoice.payment_status:
                        payment_status_choices = dict(Invoice.PAYMENT_STATUS_CHOICES)
                        payment_status_display = payment_status_choices.get(
                            corresponding_invoice.payment_status, 
                            corresponding_invoice.payment_status
                        )
                    else:
                        payment_status_display = '未付款'
                    
                    # 取得該專案的實收金額
                    project_invoice_amount = project_invoice_record.amount if project_invoice_record and project_invoice_record.amount else 0
                    
                    # 格式化日期
                    payment_received_date_str = (
                        corresponding_invoice.payment_received_date.strftime('%Y/%m/%d') 
                        if corresponding_invoice.payment_received_date else '-'
                    )
                    account_entry_date_str = (
                        corresponding_invoice.account_entry_date.strftime('%Y/%m/%d') 
                        if corresponding_invoice.account_entry_date else '-'
                    )
                    
                    # 備註
                    notes = corresponding_invoice.notes if corresponding_invoice.notes else '-'
                
                # 組裝資料行
                row_data = {
                    '案件編號': project_code,
                    '請款單號': payment.payment_number,
                    '業主': payment.owner.company_name if payment.owner else '-',
                    '收款公司': payment.company.name if payment.company else '-',
                    '請款金額': payment_project.amount,
                    '收款日': payment_received_date_str,
                    '入帳日': account_entry_date_str,
                    '收款方式': payment_method_display,
                    '實收金額': project_invoice_amount,
                    '付款狀態': payment_status_display,
                    '備註': notes,
                    '建立者': payment.created_by.username if payment.created_by else '-',
                    '建立時間': payment.created_at.strftime('%Y/%m/%d %H:%M') if payment.created_at else '-'
                }
                
                export_data.append(row_data)
        
        if not export_data:
            return Response({'message': '沒有找到可匯出的專案資料'}, 
                          status=status.HTTP_204_NO_CONTENT)
        
        # 建立Excel工作簿
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "請款單發票資料"
        
        # 設定欄位標題
        headers = ['案件編號', '請款單號', '業主', '收款公司', '請款金額', '收款日', 
                  '入帳日', '收款方式', '實收金額', '付款狀態', '備註', '建立者', '建立時間']
        
        # 寫入標題行
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(name="Microsoft JhengHei", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 寫入資料行
        for row_idx, data in enumerate(export_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = data.get(header, '')
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Microsoft JhengHei", color="000000")

                # 數字欄位右對齊
                if header in ['請款金額', '實收金額']:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = "#,##0"
                elif header in ['收款日', '入帳日', '建立時間']:
                    cell.alignment = Alignment(horizontal='center')
        
        # 自動調整欄寬
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大寬度限制為50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 準備回應
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 產生檔名
        today = timezone.now().strftime('%Y%m%d')
        date_range_str = ''
        if date_start and date_end:
            date_range_str = f"_{date_start.replace('-', '')}_{date_end.replace('-', '')}"
        elif date_start:
            date_range_str = f"_{date_start.replace('-', '')}_onwards"
        elif date_end:
            date_range_str = f"_until_{date_end.replace('-', '')}"
        
        filename = f"請款單發票資料{date_range_str}_{today}.xlsx"
        
        # 設定檔案下載標頭，支援中文檔名
        from urllib.parse import quote
        try:
            filename.encode('ascii')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        except UnicodeEncodeError:
            encoded_filename = quote(filename.encode('utf-8'))
            response['Content-Disposition'] = f'attachment; filename="export.xlsx"; filename*=UTF-8\'\'{encoded_filename}'
        
        # 儲存工作簿到回應
        workbook.save(response)
        
        return response
        
    except Exception as e:
        logger.error(f"匯出請款單發票資料時發生錯誤: {str(e)}")
        return Response({'error': f'匯出失敗: {str(e)}'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)