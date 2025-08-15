from django.core.management.base import BaseCommand
from django.core.management import call_command
from crm.models import Owner
import openpyxl


class Command(BaseCommand):
    """
    使用 fixtures 建立測試業主資料的命令
    透過 Django 的 loaddata 機制載入 JSON 檔案中的資料
    或是從 Excel 檔案匯入資料
    """

    help = "使用 fixtures 建立測試業主資料，或從 Excel 匯入"

    
    def add_arguments(self, parser):
        # 新增 excel 參數作為開關
        parser.add_argument(
            '--excel',
            action='store_true',
            help='從預設的Excel檔案載入業主資料',
        )
        parser.add_argument(
            '--force',
            action='store_true',
            help='強制載入資料，即使資料庫中已有業主資料',
        )

    def handle(self, *args, **options):
        # 取得參數
        use_excel = options.get('excel', False)
        force = options.get('force', False)
        
        # 設定固定的 Excel 檔案路徑
        excel_path = "crm/excel/06相關表格.xlsx"
        
        # 檢查是否已有業主資料
        existing_owners = Owner.objects.count()
        if existing_owners > 0 and not force:
            self.stdout.write(
                self.style.WARNING(
                    f"資料庫中已有 {existing_owners} 筆業主資料。如需重新載入，請加入 --force 參數或先清空業主資料。"
                )
            )
            return

        # 根據參數決定使用哪種方式載入資料
        if use_excel:
            self.import_from_excel(excel_path)
        else:
            # 使用 loaddata 命令載入 JSON 資料
            self.stdout.write(self.style.NOTICE("開始從 JSON 載入測試業主資料..."))
            try:
                call_command("loaddata", "test_owners_data", verbosity=1)
                loaded_count = Owner.objects.count() - existing_owners
                self.stdout.write(
                    self.style.SUCCESS(f"成功載入 {loaded_count} 筆測試業主資料")
                )
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"載入資料時發生錯誤: {str(e)}"))

    def import_from_excel(self, excel_path):
        """
        從 Excel 檔案匯入業主資料

        Args:
            excel_path (str): Excel 檔案的路徑
        """
        self.stdout.write(self.style.NOTICE(f"開始從 Excel 檔案匯入業主資料: {excel_path}"))

        try:
            # 開啟 Excel 檔案
            workbook = openpyxl.load_workbook(excel_path)
            if not workbook.sheetnames:
                self.stdout.write(self.style.ERROR("Excel 檔案中沒有工作表"))
                return

            # 使用第一張工作表
            sheet = workbook[workbook.sheetnames[0]]

            # 逐行讀取資料，跳過標題列
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 確保行數據的長度與欄位數量匹配
                row = list(row) + [None] * (7 - len(row))
                company_name, phone, fax, email, mobile, address, contact_person = row[:7]

                # 處理空值，填入預設值
                company_name = company_name or "待修改-company_name"
                phone = phone or "待修改-phone"
                fax = fax or "待修改-fax"
                email = email or "待修改-email"
                mobile = mobile or "待修改-mobile"
                address = address or "待修改-address"
                contact_person = contact_person or "待修改-contact_person"

                # 新增到 Owner 模型，先儲存以獲取自動生成的 id
                owner = Owner.objects.create(
                    company_name=company_name,
                    tax_id="",  # 暫時設為空
                    phone=phone,
                    fax=fax,
                    email=email,
                    mobile=mobile,
                    address=address,
                    contact_person=contact_person,
                )

                # 更新 tax_id 為資料庫生成的 id
                owner.tax_id = f"待修改-{owner.id}"
                owner.save(update_fields=["tax_id"])

            # 計算載入後的資料數量
            final_count = Owner.objects.count()
            self.stdout.write(
                self.style.SUCCESS(f"成功從 Excel 匯入資料，當前資料庫有 {final_count} 筆業主資料")
            )
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"從 Excel 匯入資料時發生錯誤: {str(e)}"))
