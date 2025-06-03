import csv
import os
import re
from datetime import datetime
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from django.db import transaction
from crm.models import (
    Owner,
    Category,
    Project,
    ProjectChange,
    Quotation,
    Expenditure,
    Payment,
    PaymentProject,
    Company
)
from users.models import UserProfile  # 添加這行導入 UserProfile
import openpyxl


class Command(BaseCommand):
    help = "從CSV文件導入專案資料"

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            action="store_true",
            help="從預設的Excel檔案載入專案資料",
        )
        parser.add_argument(
            "--force",
            action="store_true",
            help="強制載入資料，即使資料庫中已有專案資料",
        )
        parser.add_argument(
            "--debug",
            action="store_true",
            help="輸出除錯訊息",
        )

    def handle(self, *args, **kwargs):
        use_excel = kwargs.get("excel", False)
        force = kwargs.get("force", False)
        self.debug = kwargs.get("debug", False)

        # 設定固定的 Excel 檔案路徑
        excel_path = os.path.join("crm", "excel", "06相關表格.xlsx")

        if use_excel:
            self.stdout.write(self.style.SUCCESS(f"開始從 Excel 檔案載入專案資料，檔案位置: {excel_path}"))
            self.import_projects_from_excel(excel_path, force)
        else:
            file_path = kwargs.get("file")
            if not file_path:
                file_path = os.path.join("crm", "static", "crm", "ref", "projects.csv")

            self.stdout.write(self.style.SUCCESS(f"開始從 CSV 檔案載入專案資料，檔案位置: {file_path}"))
            self.import_projects(file_path)

    def debug_log(self, message):
        """如果開啟除錯模式，則輸出除錯訊息"""
        if self.debug:
            self.stdout.write(self.style.WARNING(f"[DEBUG] {message}"))

    def setup_default_data(self):
        """建立導入所需的預設資料"""
        # 建立預設使用者
        self.setup_default_users()

        # 建立預設類別
        self.setup_default_categories()

        # 建立預設業主
        self.setup_default_owners()
        
        # 建立預設收款公司
        self.setup_default_company()

    def setup_default_users(self):
        """建立CSV中出現的所有使用者，同時建立相應的使用者檔案"""
        users_to_create = [
            "王美淇",
            "陳俊宏",
            "王岳穎",
            "李明書",
            "楊力宇",
            "羅茗葳",
            "吳聲信",
            "陳昱仲",
            "翁佩芬",
            "蔡宗林",
            "黃柚蓁",
            "郭雅綾",
        ]

        for username in users_to_create:
            # 建立使用者
            user, user_created = User.objects.get_or_create(
                username=username, defaults={"first_name": username, "is_active": True}
            )

            # 建立或取得對應的使用者檔案
            profile, profile_created = UserProfile.objects.get_or_create(
                user=user,
                defaults={
                    "name": username,  # 使用使用者名稱作為姓名
                    "is_designer": username
                    in ["翁佩芬", "蔡宗林", "黃柚蓁", "郭雅綾"],  # 假設這些是設計師
                    "is_project_manager": username
                    in [
                        "王美淇",
                        "陳俊宏",
                        "王岳穎",
                        "李明書",
                        "楊力宇",
                        "羅茗葳",
                        "吳聲信",
                        "陳昱仲",
                    ],  # 假設這些是專案管理員
                },
            )

            if user_created:
                self.debug_log(f"建立使用者: {username}")
            if profile_created:
                self.debug_log(f"建立使用者檔案: {profile.name}")

        self.stdout.write(self.style.SUCCESS("已建立預設使用者及其檔案"))

    def setup_default_categories(self):
        """建立CSV中出現的所有類別"""
        categories_to_create = [
            {"code": "A", "description": "建築"},
            {"code": "D", "description": "其他"},
        ]

        for category in categories_to_create:
            cat, created = Category.objects.get_or_create(
                code=category["code"], defaults={"description": category["description"]}
            )
            if created:
                self.debug_log(f"建立類別: {cat.code} - {cat.description}")

        self.stdout.write(self.style.SUCCESS("已建立預設類別"))

    def check_categories(self):
        """檢查資料庫中的類別"""
        categories = Category.objects.all()
        self.stdout.write(
            self.style.SUCCESS(f"資料庫中共有 {categories.count()} 個類別:")
        )
        for cat in categories:
            self.stdout.write(f"  - {cat.code}: {cat.description}")

    def setup_default_owners(self):
        """建立CSV中出現的所有業主"""
        owner, created = Owner.objects.get_or_create(
            company_name="蘇寶華建築師事務所",
            defaults={
                "tax_id": "12345678",  # 假設稅碼，實際使用時應更新
                "phone": "04-12345678",
                "email": "info@example.com",
                "address": "彰化縣",
                "contact_person": "蘇寶華",
            },
        )
        if created:
            self.debug_log(f"建立業主: {owner.company_name}")

        self.stdout.write(self.style.SUCCESS("已建立預設業主"))

    # 在setup_default_data或handle方法中加入以下程式碼：
    def setup_default_company(self):
        """建立預設收款公司"""
        company, created = Company.objects.get_or_create(
            name="力宇設計有限公司",
            defaults={
                "name": "立信工程顧問有限公司",
                "responsible_person": "林育信",
                "tax_id": "45127101",
                "address": "500 彰化市中山路二段356巷1號",
                "phone": "04-7234988分機138",
                "fax": "04-7233033",
                "contact_person": "吳小姐"
                }
        )
        if created:
            self.debug_log(f"建立預設收款公司: {company.name}")
        self.default_company = company
        self.stdout.write(self.style.SUCCESS("已建立預設收款公司"))
    
    def parse_date(self, date_str):
        """解析日期字串"""
        if not date_str or date_str == "1999/12/31":
            return None

        try:
            # 嘗試常見日期格式
            for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"]:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None
        except Exception:
            return None

    def parse_boolean(self, bool_str):
        """將是/否轉換為布林值"""
        return bool_str == "是"

    def parse_amount(self, amount_str):
        """解析金額字串"""
        if not amount_str:
            return Decimal("0")

        # 移除非數字字元(保留小數點)
        amount_str = "".join(ch for ch in amount_str if ch.isdigit() or ch == ".")
        try:
            return Decimal(amount_str) if amount_str else Decimal("0")
        except:
            return Decimal("0")

    @transaction.atomic
    def import_projects(self, csv_file_path):
        """從CSV導入專案資料"""
        imported_count = 0
        skipped_count = 0
        error_count = 0

        # 檢查CSV檔案是否存在
        if not os.path.exists(csv_file_path):
            self.stdout.write(self.style.ERROR(f"找不到CSV檔案: {csv_file_path}"))
            return

        try:
            with open(csv_file_path, "r", encoding="utf-8-sig") as file:
                # 讀取內容檢查
                content_preview = file.read(200)
                self.debug_log(f"CSV檔案內容前200字元: {content_preview}")
                file.seek(0)  # 重置檔案指標

                reader = csv.DictReader(file)

                # 輸出CSV列名
                self.debug_log(f"CSV欄位名稱: {reader.fieldnames}")

                row_count = 0
                for row in reader:
                    row_count += 1
                    self.debug_log(
                        f"處理第 {row_count} 行: {row.get('案件類別')} - {row.get('年份')} - {row.get('案件編號')}"
                    )

                    # 跳過無效資料
                    if (
                        not row.get("案件類別")
                        or not row.get("年份")
                        or not row.get("案件編號")
                    ):
                        self.stdout.write(
                            self.style.WARNING(
                                f"跳過無效資料行 {row_count}: 缺少必要欄位"
                            )
                        )
                        skipped_count += 1
                        continue

                    try:
                        # 使用正規表達式解析類別
                        category_str = row.get("案件類別", "").strip()
                        match = re.match(r"^([A-Za-z]+)\s*(.*)$", category_str)
                        if match:
                            category_code = match.group(1)
                        else:
                            category_code = category_str

                        self.debug_log(
                            f"解析類別: '{category_str}' -> 代碼: '{category_code}'"
                        )

                        # 查詢類別
                        category = Category.objects.filter(code=category_code).first()
                        if not category:
                            self.stdout.write(
                                self.style.WARNING(
                                    f"找不到類別代碼: '{category_code}' (來自 '{category_str}')，跳過行 {row_count}"
                                )
                            )
                            skipped_count += 1
                            continue

                        # 查詢業主
                        owner_name = row.get("業主", "").strip()
                        owner = Owner.objects.filter(company_name=owner_name).first()
                        if not owner:
                            self.stdout.write(
                                self.style.WARNING(
                                    f"找不到業主: '{owner_name}'，跳過行 {row_count}"
                                )
                            )
                            skipped_count += 1
                            continue

                        # 處理使用者關聯
                        manager_name = row.get("案件負責人", "").strip()
                        manager = None
                        if manager_name:
                            # 處理多個負責人的情況 (取第一個)
                            if "\n" in manager_name:
                                manager_name = manager_name.split("\n")[0].strip()
                            manager = User.objects.filter(username=manager_name).first()
                            self.debug_log(
                                f"案件負責人: '{manager_name}' -> {'找到' if manager else '未找到'}"
                            )

                        drawing_name = row.get("繪圖", "").strip()

                        # 檢查繪圖人員是否為「無」
                        if drawing_name and drawing_name.lower() == "無":
                            drawing_name = ""

                        # 處理原本嘗試查找使用者的邏輯
                        # 現在直接使用名稱，但仍檢查使用者是否存在於系統中
                        if drawing_name:
                            user_exists = User.objects.filter(
                                username=drawing_name
                            ).exists()

                            self.debug_log(
                                f"繪圖人員: '{drawing_name}' -> {'在系統中' if user_exists else '不在系統中'}"
                            )

                        # 建立或更新Project，使用文字欄位
                        project_data = {
                            "owner": owner,
                            "name": row.get("案件名稱", ""),
                            "drawing": drawing_name,  # 直接使用文字，不再嘗試關聯使用者
                            "contact_info": row.get("聯絡方式", ""),
                            "notes": row.get("備註", ""),
                            "is_completed": self.parse_boolean(row.get("是否完成", "")),
                            "is_invoiced": self.parse_boolean(row.get("是否請款", "")),
                            "invoice_date": self.parse_date(row.get("請款日期", "")),
                            "invoice_amount": self.parse_amount(
                                row.get("請款金額", "")
                            ),
                            "payment_date": self.parse_date(row.get("收款日期", "")),
                            "invoice_issue_date": self.parse_date(
                                row.get("發票日期", "")
                            ),
                            "invoice_notes": row.get("請款備註", ""),
                            "is_paid": self.parse_boolean(row.get("是否收款", "")),
                        }

                        # 創建專案
                        project, created = Project.objects.update_or_create(
                            year=int(row.get("年份", 0)),
                            category=category,
                            project_number=row.get("案件編號", ""),
                            defaults=project_data,
                        )

                        # 設定專案負責人 (managers)
                        if manager:
                            project.managers.set([manager])
                        else:
                            project.managers.clear()  # 如果CSV中沒有指定負責人，則清空

                        # 處理變更記錄
                        changes_text = row.get("變更次數及說明", "")
                        if changes_text:
                            ProjectChange.objects.get_or_create(
                                project=project,
                                description=changes_text,
                                defaults={
                                    "created_at": datetime.now().date(),
                                    "created_by": manager,  # ProjectChange 的 created_by 仍可使用單一 manager
                                },
                            )

                        # 處理報價資訊
                        quotation_amount = self.parse_amount(row.get("報價", ""))
                        if quotation_amount > 0:
                            Quotation.objects.get_or_create(
                                project=project,
                                amount=quotation_amount,
                                defaults={"date_issued": datetime.now().date()},
                            )

                        # 處理支出資訊
                        expenditure_desc = row.get("支出", "")
                        if expenditure_desc:
                            Expenditure.objects.get_or_create(
                                project=project,
                                description=expenditure_desc,
                                defaults={
                                    "amount": Decimal("0"),
                                    "date": datetime.now().date(),
                                    "created_by": manager,
                                },
                            )

                        # 處理請款資訊 (如果已請款)
                        if self.parse_boolean(row.get("是否請款", "")):
                            payment_date = self.parse_date(row.get("請款日期", ""))
                            payment_amount = self.parse_amount(row.get("請款金額", ""))

                            if payment_date and payment_amount > 0:
                                # 建立or取得請款單
                                payment_number = f"{project.year}-{project.category.code}-{project.project_number}"
                                payment, payment_created = (
                                    Payment.objects.get_or_create(
                                        payment_number=payment_number,
                                        defaults={
                                            "owner": owner,
                                            "amount": payment_amount,
                                            "date_issued": payment_date,
                                            "paid": self.parse_boolean(
                                                row.get("是否收款", "")
                                            ),
                                            "payment_date": self.parse_date(
                                                row.get("收款日期", "")
                                            ),
                                            "notes": row.get("請款備註", ""),
                                            "created_by": manager,
                                            "company": self.default_company
                                        },
                                    )
                                )

                                # 建立請款單與專案的關聯
                                PaymentProject.objects.get_or_create(
                                    payment=payment,
                                    project=project,
                                    defaults={
                                        "amount": payment_amount,
                                        "description": row.get("請款備註", ""),
                                    },
                                )

                        action = "建立" if created else "更新"
                        self.stdout.write(
                            self.style.SUCCESS(
                                f"{action}專案: {project.year}-{category.code}-{project.project_number} {project.name}"
                            )
                        )
                        imported_count += 1

                    except Exception as e:
                        self.stdout.write(
                            self.style.ERROR(
                                f'處理行 {row_count}: {row.get("年份", "")}-{row.get("案件編號", "")} 時發生錯誤: {str(e)}'
                            )
                        )
                        error_count += 1

            self.stdout.write(
                self.style.SUCCESS(
                    f"資料導入完成! 成功: {imported_count}, 跳過: {skipped_count}, 錯誤: {error_count}"
                )
            )
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"讀取CSV檔案時發生錯誤: {str(e)}"))

    def map_excel_columns(self, sheet):
        """根據 Excel 的欄位名稱進行映射"""
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        column_mapping = {
            "案件類別": "category_code",
            "年份": "year",
            "案件編號": "project_number",
            "案件負責人": "manager_name",
            "繪圖": "drawing_name",
            "報價": "quotation",
            "案件名稱": "project_name",
            "業主": "owner_name",
            "聯絡方式": "contact_info",
            "是否完成": "is_completed",
            "收款日期": "payment_date",
            "是否請款": "is_invoiced",
            "請款日期": "invoice_date",
            "請款金額": "invoice_amount",
            "請款備註": "invoice_notes",
            "發票日期": "invoice_issue_date",
            "是否收款": "is_paid",
            "備註": "notes",
        }
        return {column_mapping.get(header, None): idx for idx, header in enumerate(headers) if header in column_mapping}

    def import_projects_from_excel(self, excel_path, force):
        """
        從 Excel 檔案匯入專案資料

        Args:
            excel_path (str): Excel 檔案的路徑
            force (bool): 是否強制覆寫資料
        """
        sheet_name = '案件清單(欄位不足)'

        try:
            workbook = openpyxl.load_workbook(excel_path)
            if sheet_name not in workbook.sheetnames:
                self.stdout.write(self.style.ERROR(f"未找到名為 '{sheet_name}' 的工作表"))
                return

            sheet = workbook[sheet_name]
            # 檢查資料庫中是否已有專案資料
            existing_projects = Project.objects.count()
            if existing_projects > 0 and not force:
                self.stdout.write(
                    self.style.WARNING(
                        f"資料庫中已有 {existing_projects} 筆專案資料。如需重新載入，請加入 --force 參數或先清空專案資料。"
                    )
                )
                return
            
            column_mapping = self.map_excel_columns(sheet)

            if not column_mapping:
                self.stdout.write(self.style.ERROR("無法找到有效的欄位名稱映射"))
                return

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    # 根據欄位名稱映射提取資料
                    data = {key: row[idx] for key, idx in column_mapping.items() if idx < len(row)}

                    # 檢查必要欄位是否存在
                    if not data.get("category_code") or not data.get("year") or not data.get("project_number"):
                        self.stdout.write(self.style.WARNING(f"跳過資料行，缺少必要欄位: {data}"))
                        continue

                    # 處理資料並新增專案
                    year = int(data["year"]) if data["year"] else None
                    project_number = str(data["project_number"]).zfill(4) if data["project_number"] else None
                    quotation = self.parse_amount(data.get("quotation"))
                    invoice_amount = self.parse_amount(data.get("invoice_amount"))
                    is_completed = self.parse_boolean(data.get("is_completed"))
                    is_invoiced = self.parse_boolean(data.get("is_invoiced"))
                    is_paid = self.parse_boolean(data.get("is_paid"))
                    payment_date = self.parse_date(data.get("payment_date"))
                    invoice_date = self.parse_date(data.get("invoice_date"))
                    invoice_issue_date = self.parse_date(data.get("invoice_issue_date"))

                    # 查找或新增類別
                    # 解析類別代碼和描述
                    category_data = data["category_code"]
                    match = re.match(r'^([A-Za-z]+)\s*(.*)$', category_data)
                    if match:
                        code = match.group(1)
                        description = match.group(2) or "此類別從excel自動新增"
                    else:
                        code = category_data
                        description = "此類別從excel自動新增"
                    
                    category, _ = Category.objects.get_or_create(
                        code=code,
                        defaults={"description": description}
                    )

                    # 查找或新增業主
                    owner_queryset = Owner.objects.filter(company_name=data["owner_name"])
                    if owner_queryset.count() > 2:
                        self.stdout.write(
                            self.style.WARNING(f"業主 '{data['owner_name']}' 超過2筆，跳過新增")
                        )
                        continue
                    elif owner_queryset.exists():
                        owner = owner_queryset.first()
                    else:
                        owner = Owner.objects.create(
                            company_name=data["owner_name"],
                            phone="待修改-phone",
                            fax="待修改-fax",
                            email="待修改-email",
                            mobile="待修改-mobile",
                            address="待修改-address",
                            contact_person="待修改-contact_person",
                        )

                    # 查找或新增案件負責人
                    managers = []
                    if data.get("manager_name"):
                        for name in re.split(r'[,\n]', data["manager_name"]):
                            name = name.strip()
                            if not name:
                                continue
                                
                            # Extract first character as first_name, rest as last_name
                            if len(name) > 1:
                                first_name = name[0]
                                last_name = name[1:]
                            else:
                                first_name = name
                                last_name = ""
                                
                            # Try to find user by first_name and last_name first
                            manager_query = User.objects.filter(first_name=first_name, last_name=last_name)
                            if manager_query.exists():
                                manager = manager_query.first()
                                user_created = False
                            else:
                                # Create new user if not found
                                manager = User.objects.create(
                                    first_name=first_name,
                                    last_name=last_name,
                                    username=name  # We'll update this immediately after
                                )
                                user_created = True
                                
                            manager.username = f"manager{manager.id}"
                            manager.set_password("password123")  # 使用 set_password 方法進行密碼哈希
                            manager.save()
                            managers.append(manager)
                            
                            # 只在創建新使用者時一併創立 UserProfile
                            if user_created:
                                UserProfile.objects.create(user=manager, name=name)
                                
                            # 輸出已新增使用者
                            self.stdout.write(
                                self.style.SUCCESS(f"新增或找到使用者: {manager.username}")
                            )

                    # 檢查是否已存在專案
                    if Project.objects.filter(year=year, category=category, project_number=project_number).exists():
                        self.stdout.write(
                            self.style.WARNING(f"跳過重複專案: {year}-{data['category_code']}-{project_number}")
                        )
                        continue

                    # 新增專案
                    project = Project.objects.create(
                        owner=owner,
                        year=year,
                        project_number=project_number,
                        name=data.get("project_name"),
                        drawing=data.get("drawing_name"),
                        contact_info=data.get("contact_info"),
                        notes=data.get("notes"),
                        is_completed=is_completed,
                        category=category,
                        is_invoiced=is_invoiced,
                        invoice_date=invoice_date,
                        invoice_amount=invoice_amount,
                        payment_date=payment_date,
                        invoice_issue_date=invoice_issue_date,
                        invoice_notes=data.get("invoice_notes"),
                        is_paid=is_paid,
                    )

                    # 設定案件負責人
                    if managers:
                        project.managers.set(managers)

                    self.stdout.write(
                        self.style.SUCCESS(f"成功新增專案: {year}-{data['category_code']}-{project_number}")
                    )

                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f"處理資料時發生錯誤: {str(e)}")
                    )

            self.stdout.write(self.style.SUCCESS("成功從 Excel 匯入專案資料"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"從 Excel 匯入資料時發生錯誤: {str(e)}"))
